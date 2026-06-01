"""
FBC Securities — Order Management System  v3.5
Virtual scrolling, table view, dealer in red, no duplicate orders on partial fills.
pyinstaller --onefile --windowed --name "FBC-Order-Manager" --clean order.py
"""

import os, sys, json, threading, uuid as _uuid, csv
import urllib.request as _urllib_req, urllib.error as _urllib_err
import urllib.request, subprocess
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
import calendar as _calendar

VERSION       = "3"
GITHUB_USER   = "Anashe-Masomeke"
GITHUB_REPO   = "order-management"
GITHUB_BRANCH = "main"
EXE_NAME      = "FBC-Order-Manager.exe"

_EXE = f"https://github.com/{GITHUB_USER}/{GITHUB_REPO}/releases/latest/download/{EXE_NAME}"
_VER = (f"https://raw.githubusercontent.com/{GITHUB_USER}/{GITHUB_REPO}/{GITHUB_BRANCH}/version.txt")

def _remote_ver():
    try:
        with urllib.request.urlopen(_VER, timeout=6) as r:
            return int(r.read().decode().strip())
    except Exception:
        return -1

def check_and_apply_update():
    rv = _remote_ver()
    if rv <= int(VERSION):  # compare against current app version
        return
    root = tk.Tk(); root.withdraw()
    ok = messagebox.askyesno("FBC Order Manager - Update Available",
        f"New version available (v{rv}).\nYour version: v{VERSION}\n\nDownload and install now?",
        icon="info")
    root.destroy()
    if not ok: return
    downloads    = os.path.join(os.path.expanduser("~"), "Downloads")
    save_dir     = downloads if os.path.isdir(downloads) else os.environ.get("TEMP", os.path.dirname(os.path.abspath(sys.argv[0])))
    new_exe_path = os.path.join(save_dir, f"FBC-Order-Manager-v{rv}.exe")
    bat_path     = os.path.join(save_dir, "_fbc_om_updater.bat")
    prog = tk.Tk(); prog.title("Downloading..."); prog.resizable(False, False)
    prog.attributes("-topmost", True)
    w, h = 420, 100
    prog.geometry(f"{w}x{h}+{(prog.winfo_screenwidth()-w)//2}+{(prog.winfo_screenheight()-h)//2}")
    tk.Label(prog, text=f"Downloading FBC Order Manager v{rv}...", font=(FONT, 10, "bold"), pady=10).pack()
    bar = ttk.Progressbar(prog, mode="indeterminate", length=360)
    bar.pack(padx=30); bar.start(12)
    lbl = tk.Label(prog, text="Starting...", font=(FONT, 8), fg="#607080"); lbl.pack(pady=4)
    prog.update(); err = [None]
    def _dl():
        try:
            dl = 0
            with urllib.request.urlopen(_EXE, timeout=180) as r:
                with open(new_exe_path, "wb") as f:
                    while True:
                        chunk = r.read(65536)
                        if not chunk: break
                        f.write(chunk); dl += len(chunk)
                        try: lbl.config(text=f"Downloaded: {dl//1024:,} KB")
                        except: pass
            if os.path.getsize(new_exe_path) < 2*1024*1024:
                os.remove(new_exe_path); raise Exception("Download incomplete.")
            with open(bat_path, "w") as f:
                f.write("\n".join(["@echo off", "ping 127.0.0.1 -n 6 > nul",
                    f'start "" "{new_exe_path}"', "ping 127.0.0.1 -n 2 > nul", 'del "%~f0"', ""]))
        except Exception as e:
            err[0] = str(e)
            for fp in [new_exe_path, bat_path]:
                try: os.remove(fp)
                except: pass
        finally:
            try: prog.after(0, prog.quit)
            except: pass
    t = threading.Thread(target=_dl, daemon=True)
    t.start(); prog.mainloop(); prog.destroy(); t.join()
    if err[0]:
        r2 = tk.Tk(); r2.withdraw()
        messagebox.showerror("Update Failed", err[0]); r2.destroy(); return
    subprocess.Popen(["cmd.exe", "/c", bat_path], creationflags=subprocess.CREATE_NO_WINDOW, close_fds=True)
    sys.exit(0)

def _pick_font():
    try:
        import tkinter.font as tkfont
        _r = tk.Tk(); _r.withdraw()
        families = tkfont.families(_r); _r.destroy()
        if "Aptos" in families: return "Aptos"
    except Exception:
        pass
    return "Segoe UI"

FONT = "Aptos"

FBC_DARK    = "#003B6F"
FBC_MID     = "#0066B3"
FBC_ACCENT  = "#00A3E0"
BG          = "#EEF2F7"
CARD_BG     = "#FFFFFF"
SEP_CLR     = "#D0DAE8"
SIDEBAR_BG  = "#001F3F"
SIDEBAR_TXT = "#B0C8E8"
WHITE       = "#FFFFFF"
TBL_HDR_BG  = "#002855"
TBL_HDR_FG  = "#90CAF9"
TBL_ROW_A   = "#FFFFFF"
TBL_ROW_B   = "#F4F7FB"
TBL_SEL     = "#E8F1FB"

S_PENDING   = ("#FFF8E7", "#B45309", "#FBC02D")
S_TAKEN     = ("#EAF4FB", "#0066B3", "#00A3E0")
S_EXECUTED  = ("#F0FFF4", "#1A6B3A", "#4CAF50")
S_PARTIAL   = ("#F3E8FF", "#6B21A8", "#A855F7")
S_CANCELLED = ("#F5F5F5", "#757575", "#9E9E9E")

AGING_OK_BG      = "#EAF7EF"; AGING_OK_CLR      = "#1A6B3A"
AGING_WARN_BG    = "#FFF8E7"; AGING_WARN_CLR    = "#B45309"
AGING_OVERDUE_BG = "#FFF0F0"; AGING_OVERDUE_CLR = "#B71C1C"

OVERDUE_DAYS  = 2
STATE_FILE    = os.path.join(os.path.expanduser("~"), ".fbc_orders.json")
SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".fbc_orders_settings.json")
OM_TABLE      = "fbc_orders"
CUSTODIANS    = ["FBC", "CABS", "CBZ", "STANBIC"]

def load_settings():
    try:
        with open(SETTINGS_FILE) as f: return json.load(f)
    except:
        return {"dealer_name": "", "supa_url": "", "supa_key": ""}

def save_settings(s):
    with open(SETTINGS_FILE, "w") as f: json.dump(s, f, indent=2)

def _supa_request(url, key, method, table, body=None, params=None):
    full = f"{url.rstrip('/')}/rest/v1/{table}"
    if params:
        full += "?" + "&".join(f"{k}={v}" for k, v in params.items())
    data = json.dumps(body).encode() if body else None
    headers = {"apikey": key, "Authorization": f"Bearer {key}",
               "Content-Type": "application/json", "Prefer": "return=representation"}
    import ssl
    _ctx = ssl.create_default_context()
    _ctx.check_hostname = False
    _ctx.verify_mode = ssl.CERT_NONE
    req = _urllib_req.Request(full, data=data, headers=headers, method=method)
    try:
        with _urllib_req.urlopen(req, timeout=15, context=_ctx) as r:
            raw = r.read().decode()
            return json.loads(raw) if raw.strip() else []
    except _urllib_err.HTTPError as e:
        raise Exception(f"HTTP {e.code}: {e.read().decode()[:200]}")

def test_sheets_connection(url, key):
    try:
        _supa_request(url, key, "GET", OM_TABLE, params={"limit": "1"})
        return True, ""
    except Exception as e:
        return False, str(e)

def _order_to_record(o):
    return {
        "id": o.get("id",""), "order_type": o.get("order_type",""),
        "client_name": o.get("client_name",""), "client_address": o.get("client_address",""),
        "csd_no": o.get("csd_no",""), "custodian": o.get("custodian",""),
        "instruction_by": o.get("instruction_by",""), "num_shares": int(o.get("num_shares",0)),
        "counter": o.get("counter",""), "limit_price": o.get("limit_price",""),
        "exchange": o.get("exchange",""), "entered_by": o.get("entered_by",""),
        "entered_datetime": o.get("entered_datetime",""), "order_date": o.get("order_date",""),
        "status": o.get("status","PENDING"), "taken_by": o.get("taken_by",""),
        "taken_datetime": o.get("taken_datetime",""), "executed_by": o.get("executed_by",""),
        "executed_datetime": o.get("executed_datetime",""),
        "shares_executed": int(o.get("shares_executed",0)),
        "execution_price": o.get("execution_price",""), "cancel_reason": o.get("cancel_reason",""),
        "notes": o.get("notes",""), "partial_of": o.get("partial_of",""),
        "amount_mode": o.get("amount_mode", "SHARES"), "total_amount": o.get("total_amount", ""),
        "tel_no": o.get("tel_no", ""),
    }

def _record_to_order(r):
    def s(k, d=""): return r.get(k, d) or d
    def si(k):
        try: return int(r.get(k, 0) or 0)
        except: return 0
    return {
        "id": s("id"), "order_type": s("order_type"), "client_name": s("client_name"),
        "client_address": s("client_address"), "csd_no": s("csd_no"),
        "custodian": s("custodian"), "instruction_by": s("instruction_by"),
        "num_shares": si("num_shares"), "counter": s("counter"), "limit_price": s("limit_price"),
        "exchange": s("exchange"), "entered_by": s("entered_by"),
        "entered_datetime": s("entered_datetime"), "order_date": s("order_date"),
        "status": s("status","PENDING"), "taken_by": s("taken_by"),
        "taken_datetime": s("taken_datetime"), "executed_by": s("executed_by"),
        "executed_datetime": s("executed_datetime"), "shares_executed": si("shares_executed"),
        "execution_price": s("execution_price"), "cancel_reason": s("cancel_reason"),
        "notes": s("notes"), "partial_of": s("partial_of"),
        "amount_mode": s("amount_mode", "SHARES"), "total_amount": s("total_amount"),
        "tel_no": s("tel_no"),
    }


class SheetsDB:
    def __init__(self, url, key):
        self.url = url; self.key = key
        self._online = False; self._lock = threading.Lock()
        self._connect()

    def _connect(self):
        try:
            _supa_request(self.url, self.key, "GET", OM_TABLE, params={"limit": "1"})
            self._online = True; print("[SupabaseDB] Connected OK")
        except Exception as e:
            self._online = False; print(f"[SupabaseDB] Offline - {e}")

    def _reconnect_if_needed(self):
        if not self._online: self._connect()

    @property
    def online(self): return self._online

    def read_all(self):
        self._reconnect_if_needed()
        if self._online:
            try:
                with self._lock:
                    rows = _supa_request(self.url, self.key, "GET", OM_TABLE,
                                         params={"order": "entered_datetime.asc", "limit": "10000"})
                orders = [_record_to_order(r) for r in rows]
                with open(STATE_FILE, "w") as f: json.dump(orders, f, indent=2)
                return orders
            except Exception as e:
                print(f"[SupabaseDB] read failed: {e}"); self._online = False
        try:
            with open(STATE_FILE) as f: return json.load(f)
        except: return []

    def append_order(self, order):
        self._reconnect_if_needed()
        if self._online:
            try:
                with self._lock:
                    _supa_request(self.url, self.key, "POST", OM_TABLE, body=_order_to_record(order))
                return True
            except Exception as e:
                print(f"[SupabaseDB] insert failed: {e}"); self._online = False
        return False

    def update_order(self, order):
        self._reconnect_if_needed()
        if self._online:
            try:
                with self._lock:
                    _supa_request(self.url, self.key, "PATCH", OM_TABLE,
                                  body=_order_to_record(order), params={"id": f"eq.{order['id']}"})
                return True
            except Exception as e:
                print(f"[SupabaseDB] update failed: {e}"); self._online = False
        return False

    def delete_order(self, order_id):
        self._reconnect_if_needed()
        if self._online:
            try:
                with self._lock:
                    _supa_request(self.url, self.key, "DELETE", OM_TABLE,
                                  params={"id": f"eq.{order_id}"})
                return True
            except Exception as e:
                print(f"[SupabaseDB] delete failed: {e}"); self._online = False
        return False


def new_id(): return str(_uuid.uuid4())[:8].upper()
def load_orders_local():
    try:
        with open(STATE_FILE) as f: return json.load(f)
    except: return []
def save_orders_local(orders):
    with open(STATE_FILE, "w") as f: json.dump(orders, f, indent=2)

def days_since_str(dt_str):
    if not dt_str: return 0
    try:
        d = datetime.fromisoformat(dt_str).date()
        return (date.today() - d).days
    except:
        try: return (date.today() - datetime.strptime(dt_str[:10], "%Y-%m-%d").date()).days
        except: return 0

def hours_since_str(dt_str):
    if not dt_str: return 0
    try:
        dt = datetime.fromisoformat(dt_str)
        return int((datetime.now() - dt).total_seconds() // 3600)
    except: return 0

def aging_info(order):
    entered = order.get("entered_datetime", "")
    if not entered: return None
    hrs = hours_since_str(entered); days = hrs // 24
    if days >= OVERDUE_DAYS: return (f"OVERDUE {days}d", AGING_OVERDUE_CLR, AGING_OVERDUE_BG)
    elif hrs >= 4: return (f"{days}d {hrs%24}h", AGING_WARN_CLR, AGING_WARN_BG)
    else: return (f"{hrs}h", AGING_OK_CLR, AGING_OK_BG)

def fmt_dt(dt_str):
    if not dt_str: return "-"
    try: return datetime.fromisoformat(dt_str).strftime("%d %b %Y  %H:%M")
    except: return dt_str[:16]

def fmt_date(d_str):
    if not d_str: return "-"
    try: return datetime.strptime(d_str, "%Y-%m-%d").strftime("%d %b %Y")
    except: return d_str

def fmt_date_short(d_str):
    if not d_str: return "-"
    try: return datetime.strptime(d_str, "%Y-%m-%d").strftime("%d/%m/%y")
    except: return d_str[:10]

def parse_date_input(s):
    s = s.strip()
    if not s: return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None


class CalendarPopup(tk.Toplevel):
    DAYS   = ["Mo","Tu","We","Th","Fr","Sa","Su"]
    MONTHS = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]

    def __init__(self, parent, initial_date=None, on_select=None):
        super().__init__(parent)
        self.on_select = on_select
        self.overrideredirect(True); self.configure(bg=FBC_DARK)
        self.attributes("-topmost", True)
        today = date.today()
        self._year  = (initial_date or today).year
        self._month = (initial_date or today).month
        self._sel   = initial_date or today
        self._build(); self._place_near(parent)
        self.bind("<FocusOut>", lambda e: self._maybe_close(e)); self.focus_set()

    def _maybe_close(self, e):
        try:
            if not self.winfo_exists(): return
            if e.widget == self: self.destroy()
        except: pass

    def _place_near(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx(); y = parent.winfo_rooty() + parent.winfo_height() + 2
        sw = parent.winfo_screenwidth(); w = self.winfo_reqwidth()
        if x + w > sw: x = sw - w - 4
        self.geometry(f"+{x}+{y}")

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        hdr = tk.Frame(self, bg=FBC_DARK, pady=4); hdr.pack(fill="x")
        tk.Button(hdr, text="<", font=(FONT, 10, "bold"), bg=FBC_DARK, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_MID,
                  command=self._prev_month).pack(side="left", padx=6)
        tk.Label(hdr, text=f"{self.MONTHS[self._month-1]}  {self._year}",
                 bg=FBC_DARK, fg=WHITE, font=(FONT, 9, "bold"), width=16,
                 anchor="center").pack(side="left", expand=True)
        tk.Button(hdr, text=">", font=(FONT, 10, "bold"), bg=FBC_DARK, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_MID,
                  command=self._next_month).pack(side="right", padx=6)
        grid = tk.Frame(self, bg="#002855", padx=6, pady=4); grid.pack()
        for ci, day in enumerate(self.DAYS):
            fg = "#FF7070" if ci >= 5 else SIDEBAR_TXT
            tk.Label(grid, text=day, bg="#002855", fg=fg,
                     font=(FONT, 8, "bold"), width=3, anchor="center").grid(row=0, column=ci, padx=1)
        cal = _calendar.monthcalendar(self._year, self._month); today = date.today()
        for ri, week in enumerate(cal):
            for ci, day in enumerate(week):
                if day == 0:
                    tk.Label(grid, text="", bg="#002855", width=3).grid(row=ri+1, column=ci, padx=1, pady=1)
                else:
                    d = date(self._year, self._month, day)
                    is_sel = (d == self._sel); is_today = (d == today)
                    if is_sel: dbg, dfg = FBC_ACCENT, WHITE
                    elif is_today: dbg, dfg = FBC_MID, WHITE
                    else:
                        dbg = "#002855"; dfg = "#FF7070" if ci >= 5 else WHITE
                    b = tk.Button(grid, text=str(day), bg=dbg, fg=dfg,
                                  font=(FONT, 8), relief="flat", cursor="hand2",
                                  width=3, activebackground=FBC_ACCENT, activeforeground=WHITE,
                                  command=lambda _d=d: self._pick(_d))
                    b.grid(row=ri+1, column=ci, padx=1, pady=1)
        foot = tk.Frame(self, bg=FBC_DARK, pady=3); foot.pack(fill="x")
        tk.Button(foot, text="Today", font=(FONT, 8), bg=FBC_MID, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_DARK,
                  command=lambda: self._pick(date.today())).pack()

    def _prev_month(self):
        if self._month == 1: self._month = 12; self._year -= 1
        else: self._month -= 1
        self._build()

    def _next_month(self):
        if self._month == 12: self._month = 1; self._year += 1
        else: self._month += 1
        self._build()

    def _pick(self, d):
        self._sel = d
        if self.on_select: self.on_select(d)
        self.destroy()


def date_entry(parent, var, bg=BG):
    frame = tk.Frame(parent, bg=bg)
    e = tk.Entry(frame, textvariable=var, font=(FONT, 9), bg="#F7FAFC", fg="#1A2B3C",
                 relief="flat", highlightbackground=SEP_CLR, highlightthickness=1,
                 width=11, cursor="hand2")
    e.pack(side="left", ipady=3)
    _popup = [None]
    def _open_cal(evt=None):
        if _popup[0] and _popup[0].winfo_exists():
            _popup[0].destroy(); _popup[0] = None; return
        initial = parse_date_input(var.get())
        def _on_sel(d):
            var.set(d.strftime("%d/%m/%Y")); _popup[0] = None
        _popup[0] = CalendarPopup(e, initial_date=initial, on_select=_on_sel)
    e.bind("<Button-1>", _open_cal)
    tk.Button(frame, text="Cal", font=(FONT, 8), bg=bg, fg=FBC_MID,
              relief="flat", cursor="hand2", activebackground=SEP_CLR,
              command=_open_cal).pack(side="left", padx=(1, 0))
    return frame


def flat_entry(parent, var, width=None, **kw):
    e = tk.Entry(parent, textvariable=var, font=(FONT, 10), bg="#F7FAFC", fg="#1A2B3C",
                 relief="flat", highlightbackground=SEP_CLR, highlightthickness=1, **kw)
    if width: e.config(width=width)
    return e

def flat_text(parent, height=5, **kw):
    return tk.Text(parent, font=(FONT, 10), bg="#F7FAFC", fg="#1A2B3C",
                   relief="flat", highlightbackground=SEP_CLR, highlightthickness=1,
                   height=height, wrap="word", **kw)

def card_frame(parent, bg=CARD_BG, **kw):
    return tk.Frame(parent, bg=bg, padx=10, pady=7,
                    highlightbackground=SEP_CLR, highlightthickness=1, **kw)

def section_lbl(parent, text, bg=BG):
    tk.Label(parent, text=text, bg=bg, fg=FBC_DARK,
             font=(FONT, 9, "bold")).pack(anchor="w", pady=(10, 2))

def make_combo(parent, var, values, **kw):
    return ttk.Combobox(parent, textvariable=var, values=values,
                        font=(FONT, 10), state="readonly", **kw)

def scrollable_body(window):
    outer = tk.Frame(window, bg=BG); outer.pack(fill="both", expand=True)
    canvas = tk.Canvas(outer, bg=BG, highlightthickness=0)
    sb = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)
    body = tk.Frame(canvas, bg=BG, padx=20, pady=6)
    cid = canvas.create_window((0, 0), window=body, anchor="nw")
    body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(cid, width=e.width))
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(-1*(e.delta//120), "units"))
    return outer, body, canvas


def _clean_csd(s): return (s or "").strip().upper().replace(" ", "")
def _clean_qty(s):
    try: return int(str(s).replace(",", "").strip())
    except: return 0
def _clean_price(s):
    try: return float(str(s).replace(",", "").strip())
    except: return 0.0
def _ticker_from_security(security): return str(security).split(".")[0].upper().strip()
def _normalize_counter(counter): return str(counter).split(".")[0].upper().strip()

def _counter_matches(csv_security, order_counter):
    ticker  = _ticker_from_security(csv_security)
    counter = _normalize_counter(order_counter)
    if ticker == counter: return True, "exact"
    short = min(len(ticker), len(counter), 4)
    if short >= 3 and ticker[:short] == counter[:short]: return True, "prefix"
    return False, None

def _name_matches(csv_name, order_name):
    cn = (csv_name or "").upper(); on = (order_name or "").upper()
    words = [w for w in cn.split() if len(w) > 2]
    if not words: return False
    return sum(1 for w in words if w in on) >= max(1, len(words) // 2)

def _direction_matches(csv_buysell, order_type):
    bs = (csv_buysell or "").strip().lower(); ot = (order_type or "").strip().lower()
    return (bs == "buy" and ot == "buy") or (bs == "sell" and ot == "sell")

def parse_matched_trades_csv(filepath):
    rows = []
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(filepath, newline="", encoding=enc) as f:
                for r in csv.DictReader(f): rows.append(dict(r))
            break
        except: rows = []
    return rows
def match_trades_to_orders(csv_rows, open_orders):
    """
    v3.6: Strict CSD-first matching.

    RULES (all must pass — no fallbacks):
    1. Only TAKEN orders are eligible. PENDING orders are NOT matched.
    2. CSD Account in CSV must exactly equal csd_no stored in the order.
    3. Security ticker must match (exchange suffix stripped from both sides).
    4. Buy/Sell direction must match order_type (Buy=BUY, Sell=SELL).

    If any rule fails the CSV row goes to unmatched. No counter-only or
    name-only fallback — those cause wrong-client matches.
    """
    from collections import defaultdict

    matched        = []
    unmatched      = []
    used_order_ids = set()

    # ── ONLY TAKEN orders are eligible ──────────────────────────────────
    eligible = [o for o in open_orders if o.get("status") in ("TAKEN", "PARTIAL")]

    # Fast CSD lookup: clean_CSD -> list of eligible orders
    csd_index = defaultdict(list)
    for o in eligible:
        key = _clean_csd(o.get("csd_no", ""))
        if key:
            csd_index[key].append(o)

    for row in csv_rows:
        csv_csd  = _clean_csd(row.get("CSD Account", "") or row.get("CSD", "") or "")
        csv_tick = _ticker_from_security(row.get("Security", ""))
        csv_dir  = (row.get("Buy/Sell") or "").strip().lower()

        best_order = None
        best_conf  = "HIGH"
        best_how   = ""

        # CSD must be present in CSV AND exist in our eligible orders
        if csv_csd and csv_csd in csd_index:
            for o in csd_index[csv_csd]:
                if o["id"] in used_order_ids:
                    continue
                # Security ticker must match
                cm, chow = _counter_matches(csv_tick, o.get("counter", ""))
                if not cm:
                    continue
                # Direction must match
                if not _direction_matches(csv_dir, o.get("order_type", "")):
                    continue
                # All three checks passed
                best_order = o
                best_how   = f"CSD (exact) + counter ({chow}) + direction"
                break

        if best_order:
            used_order_ids.add(best_order["id"])
            matched.append((best_order, [row], best_conf, best_how))
        else:
            unmatched.append(row)

    return matched, unmatched
def build_execution_from_rows(order, csv_rows):
    """
    Aggregate CSV fills for one order.
    Returns (total_qty, avg_price, status_flag, excess_qty)

    status_flag:
      "OK"            — qty fits within what was ordered
      "OVER_EXECUTED" — CSV qty exceeds ordered qty (warning shown, unchecked)

    Price: Yield column is always in cents. Divide by 100 to get ZiG.
    e.g. Yield=900 → 9.00 ZiG,  Yield=9000 → 90.00 ZiG
    """
    total_qty   = 0
    total_value = 0.0

    for r in csv_rows:
        qty = _clean_qty(r.get("Quantity", 0))
        yield_raw = _clean_price(r.get("Yield") or 0)
        market = (r.get("Market") or "").strip().upper()
        if market == "VFX" or market == "VFEX":
            price = round(yield_raw, 4)  # already USD
        else:
            price = round(yield_raw / 100, 4)  # ZSE cents → ZiG
        total_qty   += qty
        total_value += qty * price

    avg_price = round(total_value / total_qty, 4) if total_qty else 0.0

    # ── Calculate deal total including charges ───────────────────────────
    is_vfex = any((r.get("Market") or "").strip().upper() in ("VFX", "VFEX") for r in csv_rows)
    consideration = total_qty * avg_price
    order_type = order.get("order_type", "BUY").upper()
    if is_vfex:
        commission = consideration * 0.0060
        vat_comm = commission * 0.155
        vfex_levy = consideration * 0.0016
        csd_levy = consideration * 0.0005
        stamp = consideration * 0.0025
        sec_levy = consideration * 0.0016
        total_charges = commission + vat_comm + vfex_levy + csd_levy + stamp + sec_levy
    else:
        commission = consideration * 0.0092
        vat_comm = commission * 0.155
        zse_levy = consideration * 0.0010
        csd_levy = consideration * 0.0010
        stamp = consideration * 0.0025
        sec_levy = consideration * 0.0016
        inv_prot = consideration * 0.00025
        cap_gains = consideration * 0.010
        total_charges = commission + vat_comm + zse_levy + csd_levy + stamp + sec_levy + inv_prot + cap_gains
    if order_type == "BUY":
        deal_total = consideration + total_charges
    else:
        deal_total = consideration - total_charges

    ordered_qty = order.get("num_shares", 0)
    already_filled = order.get("shares_executed", 0)
    new_cumulative = already_filled + total_qty  # raw, uncapped

    # ── Check against total_amount if order was entered by amount ────────
    amount_mode = order.get("amount_mode", "SHARES")
    total_amount = order.get("total_amount", "")
    amount_filled = False
    amount_remaining = None
    if amount_mode == "AMOUNT" and total_amount:
        try:
            ordered_amount = float(str(total_amount).replace(",", ""))
            amount_filled = (deal_total >= ordered_amount * 0.995)  # 0.5% tolerance
            amount_remaining = max(0.0, ordered_amount - deal_total)
        except:
            pass

    if ordered_qty > 0 and new_cumulative > ordered_qty:
        excess = new_cumulative - ordered_qty
        return total_qty, avg_price, "OVER_EXECUTED", excess, deal_total, total_charges, amount_remaining

    return total_qty, avg_price, "OK", 0, deal_total, total_charges, amount_remaining
class SheetsSetupDialog(tk.Toplevel):
    def __init__(self, parent, settings, on_save):
        super().__init__(parent)
        self.settings = settings; self.on_save = on_save
        self.title("Supabase Database Setup")
        self.resizable(True, True); self.configure(bg=BG); self.grab_set()
        self._url = tk.StringVar(value=settings.get("supa_url",""))
        self._key = tk.StringVar(value=settings.get("supa_key",""))
        self._build()
        self.update_idletasks()
        w, h = 620, 500
        px = parent.winfo_rootx()+(parent.winfo_width()-w)//2
        py = parent.winfo_rooty()+(parent.winfo_height()-h)//2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(10,py)}")

    def _build(self):
        tk.Frame(self, bg=FBC_ACCENT, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="Connect to Supabase Database", bg=FBC_DARK, fg=WHITE,
                 font=(FONT,12,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text="All dealers use the same URL and key.",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))
        body = tk.Frame(self, bg=BG, padx=24, pady=16); body.pack(fill="both", expand=True)
        inst = tk.Frame(body, bg="#E8F1FB", padx=12, pady=10,
                        highlightbackground="#A8C4E0", highlightthickness=1)
        inst.pack(fill="x", pady=(0,16))
        tk.Label(inst,
                 text=("How to get credentials (free, 5 min):\n"
                       "1. supabase.com -> Start your project -> sign up\n"
                       "2. Create project\n3. Add fbc_orders table\n"
                       "4. Project Settings -> API -> copy URL + anon key"),
                 bg="#E8F1FB", fg=FBC_DARK, font=(FONT,9), justify="left").pack(anchor="w")
        tk.Label(body, text="Project URL", bg=BG, fg=FBC_DARK, font=(FONT,10,"bold")).pack(anchor="w")
        flat_entry(body, self._url).pack(fill="x", ipady=6, pady=(0,14))
        tk.Label(body, text="Anon / Public Key", bg=BG, fg=FBC_DARK, font=(FONT,10,"bold")).pack(anchor="w")
        flat_entry(body, self._key).pack(fill="x", ipady=6, pady=(0,10))
        self._err = tk.Label(body, text="", bg=BG, fg="#B71C1C", font=(FONT,9))
        self._err.pack(anchor="w", pady=(0,6))
        bb = tk.Frame(body, bg=BG); bb.pack(fill="x", pady=(4,0))
        tk.Button(bb, text="Skip (offline mode)", font=(FONT,9), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", activebackground=SEP_CLR,
                  command=self._skip).pack(side="right", padx=(6,0))
        tk.Button(bb, text="  Connect & Save  ", font=(FONT,10,"bold"),
                  bg=FBC_MID, fg=WHITE, relief="flat", cursor="hand2",
                  activebackground=FBC_DARK, command=self._save).pack(side="right")

    def _save(self):
        url = self._url.get().strip().rstrip("/"); key = self._key.get().strip()
        if not url or not url.startswith("http"):
            self._err.config(text="Please paste your Supabase project URL."); return
        if not key:
            self._err.config(text="Please paste your anon key."); return
        self._err.config(text="Testing connection..."); self.update()
        ok, err = test_sheets_connection(url, key)
        if not ok:
            self._err.config(text=f"{err[:120]}")
            messagebox.showerror("Connection Error", err, parent=self); return
        self.settings["supa_url"] = url; self.settings["supa_key"] = key
        save_settings(self.settings); self.on_save(url, key); self.destroy()

    def _skip(self): self.on_save("",""); self.destroy()


class LoginDialog(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FBC Order Manager - Sign In")
        self.resizable(False, False); self.configure(bg=SIDEBAR_BG)
        self.authenticated = False; self.dealer_name = ""
        self._settings = load_settings(); self._build()
        self.update_idletasks()
        w = 400; sh = self.winfo_screenheight()
        h = min(self.winfo_reqheight()+20, int(sh*0.90))
        x = (self.winfo_screenwidth()-w)//2; y = max(20,(sh-h)//2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.protocol("WM_DELETE_WINDOW", self._close)

    def _build(self):
        hdr = tk.Frame(self, bg=FBC_MID, pady=20); hdr.pack(fill="x")
        tk.Label(hdr, text="FBC Securities", bg=FBC_DARK, fg=WHITE,
                 font=(FONT,18,"bold"), padx=16, pady=6).pack()
        tk.Label(hdr, text="Order Management System", bg=FBC_MID, fg=WHITE, font=(FONT,11)).pack(pady=(2,0))
        tk.Label(hdr, text=f"v{VERSION}", bg=FBC_MID, fg="#90CAF9", font=(FONT,8)).pack()
        body = tk.Frame(self, bg=SIDEBAR_BG, padx=36, pady=20); body.pack(fill="both", expand=True)
        info = tk.Frame(body, bg="#0D2B4E", padx=14, pady=10,
                        highlightbackground=FBC_MID, highlightthickness=1)
        info.pack(fill="x", pady=(0,18))
        tk.Label(info, text="Your name is stamped on every order you enter\nand every order you take.",
                 bg="#0D2B4E", fg=SIDEBAR_TXT, font=(FONT,9), justify="left").pack(anchor="w")
        tk.Label(body, text="Your name", bg=SIDEBAR_BG, fg=SIDEBAR_TXT, font=(FONT,10,"bold")).pack(anchor="w")
        saved = self._settings.get("dealer_name","").strip()
        self._name_var = tk.StringVar(value=saved)
        self._name_entry = tk.Entry(body, textvariable=self._name_var, font=(FONT,13),
                                    bg="#0D2B4E", fg=WHITE, insertbackground=WHITE, relief="flat",
                                    highlightbackground=FBC_MID, highlightthickness=1)
        self._name_entry.pack(fill="x", ipady=9, pady=(0,4))
        self._name_entry.focus(); self._name_entry.select_range(0,"end")
        self._err = tk.Label(body, text="", bg=SIDEBAR_BG, fg="#FF6B6B", font=(FONT,9))
        self._err.pack(anchor="w", pady=(4,0))
        tk.Button(body, text="  Enter System  ", command=self._go, bg=FBC_ACCENT, fg=WHITE,
                  relief="flat", font=(FONT,11,"bold"), cursor="hand2", pady=11,
                  activebackground=FBC_MID).pack(fill="x", pady=(14,0))
        self._name_entry.bind("<Return>", lambda _: self._go())
        url = self._settings.get("supa_url","")
        tk.Label(body, text="Supabase: configured" if url else "Database: not configured yet",
                 bg=SIDEBAR_BG, fg="#6EE7B7" if url else "#FFB347",
                 font=(FONT,8)).pack(anchor="w", pady=(18,0))

    def _go(self):
        name = self._name_var.get().strip()
        if not name: self._err.config(text="Please enter your name."); return
        if len(name) < 2: self._err.config(text="Name too short."); return
        self.dealer_name = name; self.authenticated = True
        save_settings({**self._settings, "dealer_name": name}); self.destroy()

    def _close(self): self.authenticated = False; self.destroy()


class ClientLookup:
    def __init__(self, orders):
        seen = {}
        for o in orders:
            key = o.get("csd_no", "").strip()
            if not key: continue
            if key not in seen:
                seen[key] = {"client_name": o.get("client_name", ""),
                             "client_address": o.get("client_address", ""), "csd_no": key,
                             "custodian": o.get("custodian", "FBC"),
                             "instruction_by": o.get("instruction_by", "CLIENT"),
                             "tel_no": o.get("tel_no", "")}
            else:
                # Update tel_no if a newer order has it and the cached one doesn't
                if not seen[key].get("tel_no") and o.get("tel_no"):
                    seen[key]["tel_no"] = o.get("tel_no", "")
        self.clients = list(seen.values())

    def search(self, text):
        t = text.strip().lower()
        if not t: return []
        return [c for c in self.clients
                if t in c["client_name"].lower() or t in c["csd_no"].lower()][:8]


class NewOrderDialog(tk.Toplevel):
    def __init__(self, parent, dealer_name, on_saved, orders=None):
        super().__init__(parent)
        self.dealer_name = dealer_name; self.on_saved = on_saved
        self._orders_ref = orders or []
        self.title("New Order"); self.resizable(False, False)
        self.configure(bg=BG); self.grab_set(); self._build()
        self.update_idletasks()
        w, h = 660, 740
        px = parent.winfo_rootx()+(parent.winfo_width()-w)//2
        py = parent.winfo_rooty()+(parent.winfo_height()-h)//2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")
        self._canvas.yview_moveto(0)
        self._lookup = ClientLookup(self._orders_ref)

    # REPLACE WITH:
    def _show_history_btn(self, client_name, csd_no):
        self._hist_client = client_name
        self._hist_csd = csd_no
        # Create bar once, positioned above the bottom button bar
        if not hasattr(self, "_hist_bar") or not self._hist_bar.winfo_exists():
            self._hist_bar = tk.Frame(self, bg="#E8F1FB", pady=4)
            # Place it just above the btn_bar which is always the last child
            self._hist_bar.pack(fill="x", side="bottom", before=self.winfo_children()[-1])
        for w in self._hist_bar.winfo_children():
            w.destroy()
        tk.Label(self._hist_bar,
                 text=f"📋  {client_name}  —  existing client",
                 bg="#E8F1FB", fg=FBC_DARK, font=(FONT, 9, "bold")).pack(side="left", padx=12)
        tk.Button(self._hist_bar,
                  text="View Order History",
                  font=(FONT, 8, "bold"), bg=FBC_MID, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_DARK,
                  command=self._open_history).pack(side="left", padx=6)

    def _open_history(self):
        ClientHistoryDialog(self, self._hist_client,
                            self._hist_csd, self._orders_ref)
    def _build(self):
        tk.Frame(self, bg=FBC_ACCENT, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="New Order Entry", bg=FBC_DARK, fg=WHITE,
                 font=(FONT,12,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text=f"Entered by: {self.dealer_name}   .   Date: {date.today().strftime('%d %B %Y')}",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))
        _, body, self._canvas = scrollable_body(self)

        section_lbl(body, "ORDER TYPE")
        type_card = card_frame(body); type_card.pack(fill="x")
        self._order_type = tk.StringVar(value="BUY")
        row = tk.Frame(type_card, bg=CARD_BG); row.pack(anchor="w")
        self._buy_btn = tk.Button(row, text="  BUY  ", font=(FONT,12,"bold"),
                                  bg="#1A6B3A", fg=WHITE, relief="flat", cursor="hand2",
                                  activebackground="#145A32", command=lambda: self._set_type("BUY"))
        self._buy_btn.pack(side="left", padx=(0,8), ipady=6, ipadx=4)
        self._sell_btn = tk.Button(row, text="  SELL  ", font=(FONT,12,"bold"),
                                   bg=BG, fg="#607080", relief="flat", cursor="hand2",
                                   activebackground=SEP_CLR, command=lambda: self._set_type("SELL"))
        self._sell_btn.pack(side="left", ipady=6, ipadx=4)
        self._type_hint = tk.Label(type_card, text="BUY order selected",
                                   bg=CARD_BG, fg="#1A6B3A", font=(FONT,9))
        self._type_hint.pack(anchor="w", pady=(6,0))

        section_lbl(body, "EXCHANGE")
        exc_card = card_frame(body); exc_card.pack(fill="x")
        self._exchange = tk.StringVar(value="ZSE")
        er = tk.Frame(exc_card, bg=CARD_BG); er.pack(anchor="w")
        for val, txt in [("ZSE","ZSE  (ZiG)"),("VFEX","VFEX  (USD)")]:
            tk.Radiobutton(er, text=txt, variable=self._exchange, value=val,
                           bg=CARD_BG, fg=FBC_DARK, selectcolor=CARD_BG,
                           font=(FONT,10), cursor="hand2",
                           activebackground=CARD_BG).pack(side="left", padx=(0,20))

        section_lbl(body, "CLIENT DETAILS")
        cl_card = card_frame(body); cl_card.pack(fill="x")
        self._client_name = tk.StringVar(); self._client_address = tk.StringVar()
        self._csd_no = tk.StringVar(); self._tel_no = tk.StringVar()
        self._custodian = tk.StringVar(value="FBC")
        self._instruction_by = tk.StringVar(value="CLIENT")

        name_row = tk.Frame(cl_card, bg=CARD_BG); name_row.pack(fill="x", pady=2)
        tk.Label(name_row, text="Client name *", bg=CARD_BG, fg="#607080",
                 font=(FONT,9), width=18, anchor="w").pack(side="left")
        flat_entry(name_row, self._client_name).pack(side="left", fill="x", expand=True, ipady=4, padx=(4,0))

        self._suggest_frame = tk.Frame(cl_card, bg=WHITE, highlightbackground=FBC_MID, highlightthickness=1)
        self._suggest_lb = tk.Listbox(self._suggest_frame, font=(FONT,9), bg=WHITE, fg=FBC_DARK,
                                      selectbackground=FBC_MID, selectforeground=WHITE,
                                      relief="flat", activestyle="none", height=5, cursor="hand2")
        self._suggest_lb.pack(fill="x")

        def _on_name_change(*_):
            txt = self._client_name.get()
            results = self._lookup.search(txt)
            self._suggest_lb.delete(0,"end")
            if results:
                for c in results:
                    self._suggest_lb.insert("end", f"{c['client_name']}  |  {c['csd_no']}")
                self._suggest_frame.pack(fill="x", pady=(0,2))
            else:
                self._suggest_frame.pack_forget()

        def _on_select(evt):
            sel = self._suggest_lb.curselection()
            if not sel: return
            results = self._lookup.search(self._client_name.get())
            if sel[0] < len(results):
                c = results[sel[0]]
                self._client_name.set(c["client_name"])
                self._client_address.set(c["client_address"])
                self._csd_no.set(c["csd_no"])
                self._custodian.set(c["custodian"])
                self._instruction_by.set(c["instruction_by"])
                self._tel_no.set(c.get("tel_no", ""))
                self._suggest_frame.pack_forget()
                # Show history button
                self._show_history_btn(c["client_name"], c["csd_no"])

        self._client_name.trace_add("write", _on_name_change)
        self._suggest_lb.bind("<<ListboxSelect>>", _on_select)

        for lbl_txt, var in [("Client address",self._client_address),
                              ("CSD number *",self._csd_no),("Tel no.",self._tel_no)]:
            r = tk.Frame(cl_card, bg=CARD_BG); r.pack(fill="x", pady=2)
            tk.Label(r, text=lbl_txt, bg=CARD_BG, fg="#607080",
                     font=(FONT,9), width=18, anchor="w").pack(side="left")
            flat_entry(r, var).pack(side="left", fill="x", expand=True, ipady=4, padx=(4,0))
        r = tk.Frame(cl_card, bg=CARD_BG); r.pack(fill="x", pady=2)
        tk.Label(r, text="Custodian", bg=CARD_BG, fg="#607080",
                 font=(FONT,9), width=18, anchor="w").pack(side="left")
        make_combo(r, self._custodian, CUSTODIANS).pack(side="left", ipady=3, padx=(4,0))
        r = tk.Frame(cl_card, bg=CARD_BG); r.pack(fill="x", pady=2)
        tk.Label(r, text="Instruction by", bg=CARD_BG, fg="#607080",
                 font=(FONT,9), width=18, anchor="w").pack(side="left")
        make_combo(r, self._instruction_by,
                   ["CLIENT","BROKER","OTHER"]).pack(side="left", ipady=3, padx=(4,0))

        section_lbl(body, "ORDER DETAILS")
        od_card = card_frame(body); od_card.pack(fill="x")
        r = tk.Frame(od_card, bg=CARD_BG); r.pack(fill="x", pady=2)
        tk.Label(r, text="Counter *", bg=CARD_BG, fg="#607080",
                 font=(FONT,9), width=18, anchor="w").pack(side="left")
        self._counter = tk.StringVar()
        flat_entry(r, self._counter, width=24).pack(side="left", ipady=4, padx=(4,0))
        r = tk.Frame(od_card, bg=CARD_BG); r.pack(fill="x", pady=2)
        tk.Label(r, text="Limit price *", bg=CARD_BG, fg="#607080",
                 font=(FONT,9), width=18, anchor="w").pack(side="left")
        self._limit_price = tk.StringVar()
        flat_entry(r, self._limit_price, width=14).pack(side="left", ipady=4, padx=(4,0))
        tk.Frame(od_card, bg=SEP_CLR, height=1).pack(fill="x", pady=(8,5))
        tk.Label(od_card, text="Quantity - choose one method:",
                 bg=CARD_BG, fg=FBC_DARK, font=(FONT,9,"bold")).pack(anchor="w", pady=(0,4))
        self._qty_mode = tk.StringVar(value="SHARES")
        mode_row = tk.Frame(od_card, bg=CARD_BG); mode_row.pack(anchor="w", pady=(0,6))
        tk.Radiobutton(mode_row, text="Enter number of shares", variable=self._qty_mode, value="SHARES",
                       command=self._on_qty_mode, bg=CARD_BG, fg=FBC_DARK, selectcolor=CARD_BG,
                       font=(FONT,10), cursor="hand2", activebackground=CARD_BG).pack(side="left", padx=(0,20))
        tk.Radiobutton(mode_row, text="Enter total amount", variable=self._qty_mode, value="AMOUNT",
                       command=self._on_qty_mode, bg=CARD_BG, fg=FBC_DARK, selectcolor=CARD_BG,
                       font=(FONT,10), cursor="hand2", activebackground=CARD_BG).pack(side="left")
        self._shares_row = tk.Frame(od_card, bg=CARD_BG); self._shares_row.pack(fill="x", pady=2)
        tk.Label(self._shares_row, text="No. of shares *", bg=CARD_BG, fg="#607080",
                 font=(FONT,9), width=18, anchor="w").pack(side="left")
        self._num_shares = tk.StringVar()
        flat_entry(self._shares_row, self._num_shares, width=14).pack(side="left", ipady=4, padx=(4,0))
        self._amount_row = tk.Frame(od_card, bg=CARD_BG)
        tk.Label(self._amount_row, text="Total amount *", bg=CARD_BG, fg="#607080",
                 font=(FONT,9), width=18, anchor="w").pack(side="left")
        self._total_amount = tk.StringVar()
        flat_entry(self._amount_row, self._total_amount, width=14).pack(side="left", ipady=4, padx=(4,0))
        self._currency_lbl = tk.Label(self._amount_row, text="ZiG", bg=CARD_BG, fg="#607080", font=(FONT,9))
        self._currency_lbl.pack(side="left", padx=(6,0))
        self._exchange.trace_add("write", lambda *_: self._currency_lbl.config(
            text="ZiG" if self._exchange.get()=="ZSE" else "USD"))

        section_lbl(body, "NOTES (optional)")
        nc = card_frame(body); nc.pack(fill="x")
        self._notes = flat_text(nc, height=3); self._notes.pack(fill="x")
        tk.Frame(body, bg=BG, height=8).pack()
        btn_bar = tk.Frame(self, bg=BG, padx=20, pady=10); btn_bar.pack(fill="x")
        tk.Button(btn_bar, text="Cancel", font=(FONT,10), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", command=self._close,
                  activebackground=SEP_CLR).pack(side="right", padx=(6,0))
        tk.Button(btn_bar, text="  Post Order  ", font=(FONT,10,"bold"),
                  bg=FBC_MID, fg=WHITE, relief="flat", cursor="hand2",
                  activebackground=FBC_DARK, command=self._save).pack(side="right")

    def _close(self):
        try: self._canvas.unbind_all("<MouseWheel>")
        except: pass
        self.destroy()

    def _set_type(self, t):
        self._order_type.set(t)
        if t == "BUY":
            self._buy_btn.config(bg="#1A6B3A", fg=WHITE); self._sell_btn.config(bg=BG, fg="#607080")
            self._type_hint.config(text="BUY order selected", fg="#1A6B3A")
        else:
            self._sell_btn.config(bg="#B71C1C", fg=WHITE); self._buy_btn.config(bg=BG, fg="#607080")
            self._type_hint.config(text="SELL order selected", fg="#B71C1C")

    def _on_qty_mode(self):
        if self._qty_mode.get() == "SHARES":
            self._amount_row.pack_forget(); self._shares_row.pack(fill="x", pady=2)
        else:
            self._shares_row.pack_forget(); self._amount_row.pack(fill="x", pady=2)

    def _save(self):
        client = self._client_name.get().strip(); csd = self._csd_no.get().strip()
        counter = self._counter.get().strip(); limit = self._limit_price.get().strip()
        mode = self._qty_mode.get()
        if not client: messagebox.showwarning("Missing","Please enter the client name.",parent=self); return
        if not csd: messagebox.showwarning("Missing","Please enter the CSD number.",parent=self); return
        if not counter: messagebox.showwarning("Missing","Please enter a counter.",parent=self); return
        if not limit: messagebox.showwarning("Missing","Please enter the limit price.",parent=self); return
        shares = 0; total_amount = ""
        if mode == "SHARES":
            raw = self._num_shares.get().strip()
            if not raw: messagebox.showwarning("Missing", "Please enter number of shares.", parent=self); return
            try:
                shares = int(raw.replace(",", ""))
            except:
                messagebox.showwarning("Invalid", "Number of shares must be a whole number.", parent=self); return
            if shares <= 0: messagebox.showwarning("Invalid", "Shares must be greater than zero.", parent=self); return
        else:
            raw = self._total_amount.get().strip()
            if not raw: messagebox.showwarning("Missing", "Please enter the total amount.", parent=self); return
            try:
                amt = float(raw.replace(",", ""))
            except:
                messagebox.showwarning("Invalid", "Total amount must be a number.", parent=self);
                return
            total_amount = raw
            # ── Estimate shares from amount ──────────────────────────────
            try:
                price = float(limit.strip().replace(",", ""))
                if price > 0:
                    exch = self._exchange.get()
                    order_type = self._order_type.get()
                    if exch == "VFEX":
                        charge_rate = 0.0060 + 0.0016 + 0.0005 + 0.0025 + (0.0060 * 0.155) + 0.0016
                    else:
                        charge_rate = 0.0092 + 0.0010 + 0.0010 + 0.0025 + (0.0092 * 0.155) + 0.0016 + 0.00025 + 0.010
                    if order_type == "BUY":
                        est_shares = int(amt / (price * (1 + charge_rate)))
                    else:
                        est_shares = int(amt / (price * (1 - charge_rate)))
                    if est_shares > 0:
                        consideration = est_shares * price
                        if exch == "VFEX":
                            commission = consideration * 0.0060
                            vat_comm = commission * 0.155
                            vfex_levy = consideration * 0.0016
                            csd_levy = consideration * 0.0005
                            stamp = consideration * 0.0025
                            sec_levy = consideration * 0.0016
                            total_charges = commission + vat_comm + vfex_levy + csd_levy + stamp + sec_levy
                        else:
                            commission = consideration * 0.0092
                            vat_comm = commission * 0.155
                            zse_levy = consideration * 0.0010
                            csd_levy = consideration * 0.0010
                            stamp = consideration * 0.0025
                            sec_levy = consideration * 0.0016
                            inv_prot = consideration * 0.00025
                            cap_gains = consideration * 0.010
                            total_charges = commission + vat_comm + zse_levy + csd_levy + stamp + sec_levy + inv_prot + cap_gains
                        if order_type == "BUY":
                            deal_total = consideration + total_charges
                        else:
                            deal_total = consideration - total_charges
                        currency = "USD" if exch == "VFEX" else "ZiG"
                        msg = (f"Estimated shares for {currency} {amt:,.2f} @ {price}:\n\n"
                               f"  Shares:         {est_shares:,}\n"
                               f"  Consideration:  {currency} {consideration:,.2f}\n"
                               f"  Total charges:  {currency} {total_charges:,.2f}\n"
                               f"  Deal total:     {currency} {deal_total:,.2f}\n\n"
                               f"Proceed with {est_shares:,} shares?")
                        if not messagebox.askyesno("Share Estimate", msg, parent=self):
                            return
                        shares = est_shares
            except:
                pass
            # ── Duplicate detection ──────────────────────────────────────
        dupes = [o for o in self._orders_ref
                 if o.get("csd_no", "").strip().upper() == csd.strip().upper()
                 and o.get("counter", "").strip().upper() == counter.strip().upper()
                 and o.get("order_type", "") == self._order_type.get()
                 and o.get("status", "") in ("PENDING", "TAKEN", "PARTIAL")]
        if dupes:
            d = dupes[0]
            msg = (f"⚠  Possible Duplicate Order Detected\n\n"
                   f"Client {d.get('client_name', '')} already has an open "
                   f"{d.get('order_type', '')} order for {d.get('counter', '')}:\n\n"
                   f"  Order ID:  {d.get('id', '')}\n"
                   f"  Status:    {d.get('status', '')}\n"
                   f"  Shares:    {d.get('num_shares', 0):,}\n"
                   f"  Entered:   {fmt_dt(d.get('entered_datetime', ''))}\n\n"
                   f"Are you sure you want to post another order?")
            if not messagebox.askyesno("Duplicate Warning", msg, icon="warning", parent=self):
                return
        now = datetime.now()
        order = {
            "id": new_id(), "order_type": self._order_type.get(), "client_name": client,
            "client_address": self._client_address.get().strip(), "csd_no": csd,
            "custodian": self._custodian.get(), "instruction_by": self._instruction_by.get(),
            "num_shares": shares, "counter": counter.upper(), "limit_price": limit,
            "exchange": self._exchange.get(), "entered_by": self.dealer_name,
            "entered_datetime": now.isoformat(), "order_date": date.today().isoformat(),
            "status": "PENDING", "taken_by": "", "taken_datetime": "",
            "executed_by": "", "executed_datetime": "", "shares_executed": 0,
            "execution_price": "", "cancel_reason": "",
            "tel_no": self._tel_no.get().strip(),
            "notes": self._notes.get("1.0","end").strip(),
            "partial_of": "", "amount_mode": mode, "total_amount": total_amount,
        }
        try: self._canvas.unbind_all("<MouseWheel>")
        except: pass
        self.on_saved(order); self.destroy()


class ExecuteOrderDialog(tk.Toplevel):
    def __init__(self, parent, order, dealer_name, on_executed):
        super().__init__(parent)
        self.order = order; self.dealer_name = dealer_name; self.on_executed = on_executed
        self.title("Confirm Manual Execution"); self.resizable(False, False)
        self.configure(bg=BG); self.grab_set(); self._build()
        self.update_idletasks()
        w, h = 520, 480
        px = parent.winfo_rootx()+(parent.winfo_width()-w)//2
        py = parent.winfo_rooty()+(parent.winfo_height()-h)//2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    def _build(self):
        ot = self.order["order_type"]; colour = "#1A6B3A" if ot=="BUY" else "#B71C1C"
        tk.Frame(self, bg=colour, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"Confirm Execution - {ot}  {self.order['counter']}",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,11,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text=f"Client: {self.order['client_name']}  .  CSD: {self.order['csd_no']}",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))
        body = tk.Frame(self, bg=BG, padx=24, pady=14); body.pack(fill="both", expand=True)
        summ = tk.Frame(body, bg="#EAF4FB", padx=12, pady=8,
                        highlightbackground="#90CAF9", highlightthickness=1)
        summ.pack(fill="x", pady=(0,14))
        mode = self.order.get("amount_mode","SHARES")
        qty_txt = (f"Total amount: {self.order.get('total_amount','?')}  {self.order['exchange']}"
                   if mode=="AMOUNT" else f"{self.order['num_shares']:,} shares ordered")
        tk.Label(summ, text=f"{ot}  {qty_txt}  -  {self.order['counter']}  [{self.order['exchange']}]",
                 bg="#EAF4FB", fg=FBC_DARK, font=(FONT,11,"bold")).pack(anchor="w")
        already = self.order.get("shares_executed",0)
        if already:
            tk.Label(summ, text=f"Previously filled: {already:,} shares",
                     bg="#EAF4FB", fg="#6B21A8", font=(FONT,9,"italic")).pack(anchor="w")
        r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=(0,6))
        tk.Label(r, text="Shares executed *", bg=BG, fg="#607080",
                 font=(FONT,9), width=20, anchor="w").pack(side="left")
        remaining = self.order["num_shares"] - already if self.order["num_shares"] > 0 else 0
        self._shares_exec = tk.StringVar(value=str(remaining) if remaining > 0 else "")
        self._shares_entry = flat_entry(r, self._shares_exec, width=14)
        self._shares_entry.pack(side="left", ipady=4, padx=(4,0))
        r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=(0,6))
        tk.Label(r, text="Execution price *", bg=BG, fg="#607080",
                 font=(FONT,9), width=20, anchor="w").pack(side="left")
        self._exec_price = tk.StringVar(value=self.order.get("limit_price",""))
        flat_entry(r, self._exec_price, width=14).pack(side="left", ipady=4, padx=(4,0))
        note = tk.Frame(body, bg="#FFF8E7", padx=10, pady=6,
                        highlightbackground="#FBC02D", highlightthickness=1)
        note.pack(fill="x", pady=(8,4))
        tk.Label(note, text=("If shares executed < total ordered: status -> PARTIAL (updated in place).\n"
                             "If shares executed = total ordered: status -> EXECUTED."),
                 bg="#FFF8E7", fg="#B45309", font=(FONT,9), justify="left").pack(anchor="w")
        self._err = tk.Label(body, text="", bg=BG, fg="#B71C1C", font=(FONT,9))
        self._err.pack(anchor="w", pady=(4,0))
        btn_bar = tk.Frame(self, bg=BG, padx=20, pady=10); btn_bar.pack(fill="x")
        tk.Button(btn_bar, text="Cancel", font=(FONT,10), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", command=self.destroy,
                  activebackground=SEP_CLR).pack(side="right", padx=(6,0))
        tk.Button(btn_bar, text="  Confirm Execution  ", font=(FONT,10,"bold"),
                  bg="#1A6B3A", fg=WHITE, relief="flat", cursor="hand2",
                  activebackground="#145A32", command=self._execute).pack(side="right")

    def _execute(self):
        shares_raw = self._shares_exec.get().strip(); price = self._exec_price.get().strip()
        if not shares_raw: self._err.config(text="Enter shares executed."); return
        try: se = int(shares_raw.replace(",",""))
        except: self._err.config(text="Shares must be a whole number."); return
        if se <= 0: self._err.config(text="Shares must be greater than zero."); return
        total_ordered = self.order["num_shares"]; already = self.order.get("shares_executed",0)
        cumulative = already + se
        if total_ordered > 0 and cumulative > total_ordered:
            self._err.config(text=f"Total filled ({cumulative:,}) would exceed ordered ({total_ordered:,})."); return
        if not price: self._err.config(text="Enter execution price."); return
        now = datetime.now().isoformat()
        self.order["shares_executed"] = cumulative; self.order["execution_price"] = price
        self.order["executed_by"] = self.dealer_name; self.order["executed_datetime"] = now
        is_full = (total_ordered == 0 or cumulative >= total_ordered)
        self.order["status"] = "EXECUTED" if is_full else "PARTIAL"
        if not self.order.get("taken_by"):
            self.order["taken_by"] = self.dealer_name; self.order["taken_datetime"] = now
        self.on_executed(self.order); self.destroy()


class CancelOrderDialog(tk.Toplevel):
    def __init__(self, parent, order, dealer_name, on_cancelled):
        super().__init__(parent)
        self.order = order; self.dealer_name = dealer_name; self.on_cancelled = on_cancelled
        self.title("Cancel Order"); self.resizable(False, False)
        self.configure(bg=BG); self.grab_set(); self._build()
        self.update_idletasks()
        w, h = 460, 360
        px = parent.winfo_rootx()+(parent.winfo_width()-w)//2
        py = parent.winfo_rooty()+(parent.winfo_height()-h)//2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    def _build(self):
        tk.Frame(self, bg="#B71C1C", height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"Cancel Order - {self.order['counter']}",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,11,"bold")).pack(padx=16, anchor="w")
        shares_txt = (f"{self.order['num_shares']:,} shares" if self.order["num_shares"]>0
                      else f"Amount: {self.order.get('total_amount','?')}")
        tk.Label(hdr, text=f"{self.order['order_type']}  {shares_txt}  .  {self.order['client_name']}",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))
        body = tk.Frame(self, bg=BG, padx=24, pady=14); body.pack(fill="both", expand=True)
        warn = tk.Frame(body, bg="#FFF0F0", padx=12, pady=8,
                        highlightbackground="#FFCDD2", highlightthickness=1)
        warn.pack(fill="x", pady=(0,14))
        tk.Label(warn, text="This will permanently cancel the order.\nReason is recorded in the audit trail.",
                 bg="#FFF0F0", fg="#B71C1C", font=(FONT,9), justify="left").pack(anchor="w")
        tk.Label(body, text="Cancellation reason *", bg=BG, fg=FBC_DARK,
                 font=(FONT,10,"bold")).pack(anchor="w")
        qf = tk.Frame(body, bg=BG); qf.pack(fill="x", pady=(4,6))
        for reason in ["Client withdrew instruction","Limit not reached","Duplicate order","Other"]:
            tk.Button(qf, text=reason, font=(FONT,8), bg=BG, fg=FBC_MID,
                      relief="flat", cursor="hand2", activebackground=SEP_CLR,
                      command=lambda r=reason: self._set_reason(r)).pack(side="left", padx=(0,6))
        self._reason_txt = flat_text(body, height=4); self._reason_txt.pack(fill="x")
        self._err = tk.Label(body, text="", bg=BG, fg="#B71C1C", font=(FONT,9))
        self._err.pack(anchor="w", pady=(4,0))
        btn_bar = tk.Frame(self, bg=BG, padx=20, pady=10); btn_bar.pack(fill="x")
        tk.Button(btn_bar, text="Back", font=(FONT,10), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", command=self.destroy,
                  activebackground=SEP_CLR).pack(side="right", padx=(6,0))
        tk.Button(btn_bar, text="  Confirm Cancel  ", font=(FONT,10,"bold"),
                  bg="#B71C1C", fg=WHITE, relief="flat", cursor="hand2",
                  activebackground="#8B0000", command=self._cancel).pack(side="right")

    def _set_reason(self, r):
        self._reason_txt.delete("1.0","end"); self._reason_txt.insert("1.0",r)

    def _cancel(self):
        reason = self._reason_txt.get("1.0","end").strip()
        if not reason: self._err.config(text="Please provide a cancellation reason."); return
        self.order["status"] = "CANCELLED"; self.order["cancel_reason"] = reason
        self.on_cancelled(self.order); self.destroy()


class OrderDetailDialog(tk.Toplevel):
    def __init__(self, parent, order):
        super().__init__(parent)
        self.order = order
        self.title(f"Order Detail - {order['id']}")
        self.resizable(True, True); self.configure(bg=BG); self.grab_set(); self._build()
        self.update_idletasks()
        w, h = 540, 620
        px = parent.winfo_rootx()+(parent.winfo_width()-w)//2
        py = parent.winfo_rooty()+(parent.winfo_height()-h)//2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    def _build(self):
        ot = self.order["order_type"]; col = "#1A6B3A" if ot=="BUY" else "#B71C1C"
        tk.Frame(self, bg=col, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"Order {self.order['id']}  -  {ot}  {self.order['counter']}",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,12,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text=f"Date: {fmt_date(self.order.get('order_date',''))}   .   Entered: {fmt_dt(self.order.get('entered_datetime',''))}",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))
        _, body, _ = scrollable_body(self)
        def divider(): tk.Frame(body, bg=SEP_CLR, height=1).pack(fill="x", pady=6)
        def detail_row(label, value, bold=False):
            r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=2)
            tk.Label(r, text=label, bg=BG, fg="#607080",
                     font=(FONT,9), width=22, anchor="w").pack(side="left")
            tk.Label(r, text=str(value) if value else "-", bg=BG, fg=FBC_DARK,
                     font=(FONT,9,"bold" if bold else "normal"),
                     anchor="w", wraplength=280, justify="left").pack(side="left", fill="x", expand=True)
        o = self.order; mode = o.get("amount_mode","SHARES")
        tk.Label(body, text="ORDER", bg=BG, fg=FBC_MID, font=(FONT,8,"bold")).pack(anchor="w", pady=(8,3))
        detail_row("Order ID",o["id"],bold=True); detail_row("Type",o["order_type"],bold=True)
        detail_row("Exchange",o["exchange"]); detail_row("Counter",o["counter"],bold=True)
        detail_row("Order date",fmt_date(o.get("order_date","")))
        if mode=="AMOUNT":
            detail_row("Total amount",f"{o.get('total_amount','?')}  {o['exchange']}")
            detail_row("Shares ordered",f"{o['num_shares']:,}" if o['num_shares']>0 else "TBD")
        else:
            detail_row("Shares ordered",f"{o['num_shares']:,}")
        detail_row("Limit price",o["limit_price"])
        if o.get("shares_executed",0) > 0:
            remaining = o["num_shares"] - o["shares_executed"]
            detail_row("Shares filled",f"{o['shares_executed']:,}",bold=True)
            detail_row("Shares remaining",f"{remaining:,}" if remaining>0 else "-",bold=(remaining>0))
            detail_row("Execution price",o["execution_price"])
        divider()
        tk.Label(body, text="CLIENT", bg=BG, fg=FBC_MID, font=(FONT,8,"bold")).pack(anchor="w", pady=(0,3))
        detail_row("Client name",o["client_name"],bold=True)
        detail_row("Client address",o["client_address"]); detail_row("CSD number",o["csd_no"])
        detail_row("Custodian",o["custodian"]); detail_row("Instruction by",o["instruction_by"])
        divider()
        tk.Label(body, text="WORKFLOW", bg=BG, fg=FBC_MID, font=(FONT,8,"bold")).pack(anchor="w", pady=(0,3))
        detail_row("Entered by",o["entered_by"]); detail_row("Entered at",fmt_dt(o["entered_datetime"]))
        detail_row("Status",o["status"],bold=True); detail_row("Taken by",o["taken_by"])
        detail_row("Taken at",fmt_dt(o["taken_datetime"])); detail_row("Executed by",o["executed_by"])
        detail_row("Executed at",fmt_dt(o["executed_datetime"]))
        if o.get("cancel_reason"):
            divider()
            tk.Label(body, text="CANCELLATION", bg=BG, fg="#B71C1C", font=(FONT,8,"bold")).pack(anchor="w", pady=(0,3))
            detail_row("Cancel reason",o["cancel_reason"])
        if o.get("notes"):
            divider()
            tk.Label(body, text="NOTES", bg=BG, fg=FBC_MID, font=(FONT,8,"bold")).pack(anchor="w", pady=(0,3))
            tk.Label(body, text=o["notes"], bg=BG, fg=FBC_DARK,
                     font=(FONT,9), wraplength=460, justify="left").pack(anchor="w")
        tk.Frame(body, bg=BG, height=10).pack()
        tk.Button(body, text="  Close  ", font=(FONT,10), bg=FBC_MID, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_DARK,
                  command=self.destroy, pady=5).pack(pady=(0,10))


class MatchedTradesDialog(tk.Toplevel):
    def __init__(self, parent, orders, dealer_name, on_bulk_execute):
        super().__init__(parent)
        self.all_orders = orders; self.dealer_name = dealer_name
        self.on_bulk_execute = on_bulk_execute
        self.title("Matched Trades - Auto Reconciliation")
        self.resizable(True, True); self.configure(bg=BG); self.grab_set()
        self._files=[]; self._matched=[]; self._unmatched=[]; self._check_vars={}
        self._build()
        self.update_idletasks()
        w, h = 920, 720
        px = parent.winfo_rootx()+(parent.winfo_width()-w)//2
        py = parent.winfo_rooty()+(parent.winfo_height()-h)//2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    def _build(self):
        tk.Frame(self, bg=FBC_ACCENT, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="Matched Trades - Post-Market Auto Reconciliation",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,12,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text=("Upload ZSE/VFEX matched-trades CSV(s) after market close.\n"
                             "Each matched order updated IN PLACE. No duplicate orders created."),
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))
        pick_bar = tk.Frame(self, bg=BG, padx=16, pady=10); pick_bar.pack(fill="x")
        tk.Button(pick_bar, text="  Browse for CSV file(s)  ",
                  font=(FONT,10,"bold"), bg=FBC_MID, fg=WHITE, relief="flat", cursor="hand2",
                  activebackground=FBC_DARK, command=self._pick_files).pack(side="left")
        self._file_lbl = tk.Label(pick_bar, text="No files selected",
                                  bg=BG, fg="#8096B0", font=(FONT,9))
        self._file_lbl.pack(side="left", padx=12)
        tk.Button(pick_bar, text="  Run Matching  ",
                  font=(FONT,10,"bold"), bg="#1A6B3A", fg=WHITE, relief="flat", cursor="hand2",
                  activebackground="#145A32", command=self._run_match).pack(side="right")
        tk.Frame(self, bg=SEP_CLR, height=1).pack(fill="x")
        _, self._body, self._scroll_canvas = scrollable_body(self)
        self._confirm_bar = tk.Frame(self, bg=BG, padx=16, pady=10); self._confirm_bar.pack(fill="x")
        tk.Button(self._confirm_bar, text="Close", font=(FONT,10), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", activebackground=SEP_CLR,
                  command=self.destroy).pack(side="right", padx=(8,0))
        self._confirm_btn = tk.Button(self._confirm_bar, text="  Confirm & Update Selected  ",
                  font=(FONT,10,"bold"), bg="#1A6B3A", fg=WHITE, relief="flat", cursor="hand2",
                  activebackground="#145A32", state="disabled", command=self._confirm_execute)
        self._confirm_btn.pack(side="right")
        self._summary_lbl = tk.Label(self._confirm_bar, text="", bg=BG, fg=FBC_DARK, font=(FONT,9))
        self._summary_lbl.pack(side="left")
        tk.Label(self._body, text="Select matched-trades CSV file(s), then click Run Matching.",
                 bg=BG, fg="#8096B0", font=(FONT,11), pady=60).pack()

    def _pick_files(self):
        paths = filedialog.askopenfilenames(
            title="Select matched-trades CSV file(s)",
            filetypes=[("CSV files","*.csv"),("All files","*.*")], parent=self)
        if paths:
            self._files = list(paths)
            self._file_lbl.config(text=",  ".join(os.path.basename(p) for p in self._files), fg=FBC_DARK)

    def _run_match(self):
        if not self._files:
            messagebox.showwarning("No files","Please select at least one CSV file.",parent=self); return
        all_csv_rows = []
        for fp in self._files: all_csv_rows.extend(parse_matched_trades_csv(fp))
        if not all_csv_rows:
            messagebox.showerror("Empty","No trades found in the selected file(s).",parent=self); return
        self._matched, self._unmatched = match_trades_to_orders(all_csv_rows, self.all_orders)
        self._check_vars = {}; self._render_results()

    def _render_results(self):
        for w in self._body.winfo_children(): w.destroy()
        total_csv = sum(len(rows) for _,rows,_,_ in self._matched)+len(self._unmatched)
        n_matched = len(self._matched); n_unmatched = len(self._unmatched)
        banner_bg = "#EAF7EF" if n_unmatched==0 else "#FFF8E7"
        banner_fg = "#1A6B3A" if n_unmatched==0 else "#B45309"
        banner = tk.Frame(self._body, bg=banner_bg, padx=14, pady=8,
                          highlightbackground=banner_fg, highlightthickness=1)
        banner.pack(fill="x", pady=(4,8))
        unmatch_note = f"   .   {n_unmatched} CSV row(s) could not be matched." if n_unmatched else ""
        tk.Label(banner,
                 text=f"{'OK' if n_unmatched==0 else 'WARN'}  {n_matched} order(s) matched from {total_csv} CSV record(s).{unmatch_note}",
                 bg=banner_bg, fg=banner_fg, font=(FONT,10,"bold")).pack(anchor="w")
        tk.Label(banner, text="Each order updated IN PLACE. Partial fills keep same dealer, status -> PARTIAL.",
                 bg=banner_bg, fg=banner_fg, font=(FONT,9)).pack(anchor="w")
        if self._matched:
            tk.Label(self._body, text="PROPOSED UPDATES",
                     bg=BG, fg=FBC_MID, font=(FONT,9,"bold")).pack(anchor="w", pady=(4,3))
        for order, csv_rows, conf, how in self._matched:
            # REPLACE WITH:
            qty, avg_price, exec_flag, excess_qty, deal_total, total_charges, amount_remaining = build_execution_from_rows(order, csv_rows)
            oid = order["id"];
            ordered_qty = order.get("num_shares", 0)
            already_filled = order.get("shares_executed", 0)
            is_over = (exec_flag == "OVER_EXECUTED")
            new_total_fill = (already_filled + qty) if not is_over else ordered_qty
            is_partial = (ordered_qty > 0 and new_total_fill < ordered_qty)
            remaining_qty = max(0, ordered_qty - new_total_fill) if is_partial else 0
            held_by = order.get("taken_by", "") or order.get("executed_by", "")
            # Over-executed orders are UNCHECKED by default — dealer must review manually
            var = tk.BooleanVar(value=(False if is_over else conf in ("HIGH", "MEDIUM")))
            self._check_vars[oid] = var
            conf_col = {"HIGH":"#1A6B3A","MEDIUM":"#B45309","LOW":"#B71C1C"}.get(conf,"#607080")
            conf_bg  = {"HIGH":"#F0FFF4","MEDIUM":"#FFF8E7","LOW":"#FFF0F0"}.get(conf,BG)
            border   = {"HIGH":"#4CAF50","MEDIUM":"#FBC02D","LOW":"#EF9A9A"}.get(conf,SEP_CLR)
            card = tk.Frame(self._body, bg=conf_bg, padx=10, pady=8,
                            highlightbackground=border, highlightthickness=1)
            card.pack(fill="x", pady=2)
            tk.Checkbutton(card, variable=var, bg=conf_bg, activebackground=conf_bg,
                           cursor="hand2", command=self._update_summary).pack(side="left", anchor="n", pady=(2,0))
            info = tk.Frame(card, bg=conf_bg); info.pack(side="left", fill="x", expand=True, padx=(6,0))
            ot = order["order_type"]; badge_col = "#1A6B3A" if ot=="BUY" else "#B71C1C"
            r1 = tk.Frame(info, bg=conf_bg); r1.pack(anchor="w")
            tk.Label(r1, text=f" {ot} ", bg=badge_col, fg=WHITE,
                     font=(FONT,9,"bold"), padx=6, pady=1).pack(side="left")
            tk.Label(r1, text=f"  {order['counter']}  .  #{oid}  .  {order['client_name']}",
                     bg=conf_bg, fg=FBC_DARK, font=(FONT,10,"bold")).pack(side="left")
            if held_by:
                tk.Label(r1, text=f"  {held_by}", bg=conf_bg,
                         fg="#C0392B", font=(FONT,9,"bold")).pack(side="left", padx=(8,0))
                # Show amount or shares depending on mode
                order_mode = order.get("amount_mode", "SHARES")
                order_total_amt = order.get("total_amount", "")
                order_exch = order.get("exchange", "ZSE")
                order_currency = "USD" if order_exch == "VFEX" else "ZiG"
                if order_mode == "AMOUNT" and order_total_amt:
                    try:
                        ordered_display = f"Ordered: {order_currency} {float(str(order_total_amt).replace(',', '')):.2f}"
                    except:
                        ordered_display = f"Ordered: {order_total_amt}"
                else:
                    ordered_display = f"Ordered: {ordered_qty:,} shares"
                already_filled_txt = (f"   .   Previously filled: {already_filled:,} sh" if already_filled else "")
                tk.Label(info,
                         text=f"CSD: {order.get('csd_no', '')}   .   {ordered_display}{already_filled_txt}",
                         bg=conf_bg, fg="#607080", font=(FONT, 8)).pack(anchor="w", pady=(1, 0))
            fill_label = "PARTIAL FILL" if is_partial else "FULL FILL"
            fill_col   = "#6B21A8" if is_partial else "#1A6B3A"
            currency = "USD" if any((r.get("Market", "") or "").upper() in ("VFX", "VFEX") for r in csv_rows) else "ZiG"
            charges_line = f"  |  charges: {currency} {total_charges:,.2f}  |  deal total: {currency} {deal_total:,.2f}"
            amount_line = ""
            if amount_remaining is not None:
                amount_line = f"  →  {currency} {amount_remaining:,.2f} remaining" if amount_remaining > 0.5 else "  →  AMOUNT FULLY FILLED"
            tk.Label(info,
                     text=f"-> [{fill_label}]  +{qty:,} shares @ {avg_price}  =  {new_total_fill:,} total filled{charges_line}{amount_line}",
                     bg=conf_bg, fg=fill_col, font=(FONT,10,"bold")).pack(anchor="w", pady=(4,0))
            # REPLACE WITH:
            if is_over:
                over_frame = tk.Frame(info, bg="#FFF0F0", padx=8, pady=5,
                                      highlightbackground="#B71C1C", highlightthickness=2)
                over_frame.pack(fill="x", pady=(4, 2))
                tk.Label(over_frame,
                         text=f"🚨  OVER-EXECUTION WARNING",
                         bg="#FFF0F0", fg="#B71C1C", font=(FONT, 9, "bold")).pack(anchor="w")
                tk.Label(over_frame,
                         text=(f"CSV matched {already_filled + qty:,} shares total "
                               f"but order is only {ordered_qty:,} shares.\n"
                               f"Excess: {excess_qty:,} shares. "
                               f"Verify with broker before confirming.\n"
                               f"System will cap execution at {ordered_qty:,} shares if confirmed."),
                         bg="#FFF0F0", fg="#B71C1C", font=(FONT, 8), justify="left").pack(anchor="w")
            elif is_partial:
                tk.Label(info,
                         text=f"{remaining_qty:,} shares remain . Order stays PARTIAL with {held_by or 'same dealer'}",
                         bg=conf_bg, fg="#6B21A8", font=(FONT, 8, "italic")).pack(anchor="w", pady=(1, 0))
            tk.Label(info, text=f"Match: {conf}   .   {how}",
                     bg=conf_bg, fg=conf_col, font=(FONT, 8)).pack(anchor="w", pady=(1, 0))
        if self._unmatched:
            tk.Frame(self._body, bg=SEP_CLR, height=1).pack(fill="x", pady=(10,3))
            tk.Label(self._body, text="UNMATCHED CSV ROWS - manual review needed",
                     bg=BG, fg="#B45309", font=(FONT,9,"bold")).pack(anchor="w", pady=(0,3))
            for row in self._unmatched:
                ucard = tk.Frame(self._body, bg="#FFF8E7", padx=10, pady=5,
                                 highlightbackground="#FBC02D", highlightthickness=1)
                ucard.pack(fill="x", pady=2)
                qty = _clean_qty(row.get("Quantity",0))
                price = _clean_price(row.get("Yield") or row.get("Price") or 0)
                tk.Label(ucard,
                         text=(f"{row.get('Buy/Sell','')}  {row.get('Security','')}  .  "
                               f"{qty:,} shares @ {price}  .  {row.get('Name','')}  .  "
                               f"CSD: {row.get('CSD Account', row.get('CSD',''))}"),
                         bg="#FFF8E7", fg="#B45309", font=(FONT,9)).pack(anchor="w")
        tk.Frame(self._body, bg=BG, height=14).pack()
        self._update_summary()

    def _update_summary(self):
        n = sum(1 for var in self._check_vars.values() if var.get())
        self._summary_lbl.config(text=f"{n} order(s) selected for update")
        self._confirm_btn.config(state="normal" if n>0 else "disabled")

    def _confirm_execute(self):
        selected_ids = {oid for oid,var in self._check_vars.items() if var.get()}
        now = datetime.now().isoformat(); updates=[]
        for order, csv_rows, conf, how in self._matched:
            if order["id"] not in selected_ids: continue
            # REPLACE WITH:
            qty, avg_price, exec_flag, excess_qty, deal_total, total_charges, amount_remaining = build_execution_from_rows(
                order, csv_rows)
            ordered_qty = order.get("num_shares", 0)
            already_filled = order.get("shares_executed", 0)
            is_over = (exec_flag == "OVER_EXECUTED")
            # Cap at ordered qty — never allow over-execution in the database
            new_total_fill = ordered_qty if is_over else (already_filled + qty)
            is_partial = (ordered_qty > 0 and new_total_fill < ordered_qty)
            order["shares_executed"] = new_total_fill; order["execution_price"] = str(avg_price)
            order["executed_by"] = f"AUTO - {self.dealer_name} (matched trades)"
            order["executed_datetime"] = now
            if not order.get("taken_by"):
                order["taken_by"] = self.dealer_name; order["taken_datetime"] = now
            order["status"] = "PARTIAL" if is_partial else "EXECUTED"
            # REPLACE WITH:
            over_note = (f"  ⚠ OVER-EXECUTION: CSV showed {already_filled + qty:,}, "
                         f"capped at {ordered_qty:,} (excess {excess_qty:,} sh)" if is_over else "")
            currency = "USD" if any((r.get("Market", "") or "").upper() in ("VFX", "VFEX") for r in csv_rows) else "ZiG"
            note_line = (f"[{now[:10]}] Matched trade: +{qty:,} sh @ {avg_price}"
                         + (f"  ({new_total_fill:,}/{ordered_qty:,} filled)" if ordered_qty else "")
                         + f"  deal total: {currency} {deal_total:,.2f}"
                         + over_note)
            # ── Amount-mode: override status based on deal total vs ordered amount ──
            amount_mode_o = order.get("amount_mode", "SHARES")
            total_amount_str = order.get("total_amount", "")
            if amount_mode_o == "AMOUNT" and total_amount_str:
                try:
                    ordered_amt = float(str(total_amount_str).replace(",", ""))
                    if deal_total >= ordered_amt * 0.995:
                        order["status"] = "EXECUTED"
                        note_line += (f"  | AMOUNT FULLY FILLED"
                                      f" ({currency} {deal_total:,.2f}"
                                      f" / {ordered_amt:,.2f})")
                    else:
                        remaining_amt = ordered_amt - deal_total
                        order["status"] = "PARTIAL"
                        note_line += (f"  | AMOUNT PARTIAL:"
                                      f" {currency} {remaining_amt:,.2f} remaining"
                                      f" of {ordered_amt:,.2f}")
                except:
                    pass

            existing = order.get("notes", "")
            order["notes"] = (existing + "\n" + note_line).strip() if existing else note_line
            updates.append(order)
        if not updates:
            messagebox.showinfo("Nothing to do","No orders were selected.",parent=self); return
        partial_count = sum(1 for o in updates if o["status"]=="PARTIAL")
        full_count    = sum(1 for o in updates if o["status"]=="EXECUTED")
        details = []
        if full_count: details.append(f"{full_count} fully executed")
        if partial_count: details.append(f"{partial_count} partially filled (remain with same dealer)")
        ok = messagebox.askyesno("Confirm Bulk Update",
            f"Update {len(updates)} order(s)?\n\n"+"\n".join(f"  - {d}" for d in details)+
            "\n\nEach order updated in place. Cannot be undone.", parent=self)
        if not ok: return
        self.on_bulk_execute(updates)
        summary = f"{full_count} order(s) EXECUTED"
        if partial_count: summary += f",  {partial_count} marked PARTIAL (same order, same dealer)"
        messagebox.showinfo("Done", f"{summary}.\nThe order list has been updated.", parent=self)
        self.destroy()


class EditOrderDialog(tk.Toplevel):
    def __init__(self, parent, order, on_saved):
        super().__init__(parent)
        self.order = order; self.on_saved = on_saved
        self.title(f"Edit Order - {order['id']}")
        self.resizable(False, False); self.configure(bg=BG); self.grab_set(); self._build()
        self.update_idletasks()
        w, h = 520, 380
        px = parent.winfo_rootx()+(parent.winfo_width()-w)//2
        py = parent.winfo_rooty()+(parent.winfo_height()-h)//2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    def _build(self):
        ot = self.order["order_type"]; col = "#1A6B3A" if ot=="BUY" else "#B71C1C"
        tk.Frame(self, bg=col, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"Edit Order - {ot}  {self.order['counter']}",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,11,"bold")).pack(padx=16, anchor="w")
        body = tk.Frame(self, bg=BG, padx=24, pady=14); body.pack(fill="both", expand=True)
        info = tk.Frame(body, bg="#FFF8E7", padx=12, pady=8,
                        highlightbackground="#FBC02D", highlightthickness=1)
        info.pack(fill="x", pady=(0,14))
        tk.Label(info, text="Counter, shares, limit price and notes are editable.",
                 bg="#FFF8E7", fg="#B45309", font=(FONT,9), justify="left").pack(anchor="w")
        def frow(label, var):
            r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=3)
            tk.Label(r, text=label, bg=BG, fg="#607080", font=(FONT,9), width=18, anchor="w").pack(side="left")
            flat_entry(r, var, width=20).pack(side="left", ipady=4, padx=(4,0))
        self._counter     = tk.StringVar(value=self.order.get("counter",""))
        self._limit_price = tk.StringVar(value=self.order.get("limit_price",""))
        self._num_shares  = tk.StringVar(value=str(self.order["num_shares"]) if self.order.get("num_shares") else "")
        self._notes_var   = tk.StringVar(value=self.order.get("notes",""))
        frow("Counter *", self._counter); frow("Limit price *", self._limit_price)
        frow("No. of shares *", self._num_shares); frow("Notes", self._notes_var)
        self._err = tk.Label(body, text="", bg=BG, fg="#B71C1C", font=(FONT,9))
        self._err.pack(anchor="w", pady=(6,0))
        btn_bar = tk.Frame(self, bg=BG, padx=20, pady=10); btn_bar.pack(fill="x")
        tk.Button(btn_bar, text="Cancel", font=(FONT,10), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", command=self.destroy,
                  activebackground=SEP_CLR).pack(side="right", padx=(6,0))
        tk.Button(btn_bar, text="  Save Changes  ", font=(FONT,10,"bold"),
                  bg=FBC_MID, fg=WHITE, relief="flat", cursor="hand2",
                  activebackground=FBC_DARK, command=self._save).pack(side="right")

    def _save(self):
        counter = self._counter.get().strip().upper(); limit = self._limit_price.get().strip()
        shares_raw = self._num_shares.get().strip()
        if not counter: self._err.config(text="Counter is required."); return
        if not limit: self._err.config(text="Limit price is required."); return
        if shares_raw:
            try:
                shares = int(shares_raw.replace(",",""))
                if shares <= 0: raise ValueError
            except: self._err.config(text="Shares must be a positive whole number."); return
            self.order["num_shares"] = shares
        self.order["counter"] = counter; self.order["limit_price"] = limit
        self.order["notes"] = self._notes_var.get().strip()
        self.on_saved(self.order); self.destroy()
class AgingDialog(tk.Toplevel):
    """
    Dealer performance + order aging analysis with visual charts.
    """
    def __init__(self, parent, orders, dealer_name):
        super().__init__(parent)
        self.all_orders  = orders
        self.dealer_name = dealer_name
        self.title("Aging & Dealer Performance Analysis")
        self.resizable(True, True); self.configure(bg=BG); self.grab_set()
        self._build()
        self.update_idletasks()
        w, h = 1100, 780
        px = parent.winfo_rootx() + (parent.winfo_width()  - w) // 2
        py = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    @staticmethod
    def _hours_since(dt_str):
        if not dt_str: return 0
        try: return max(0,(datetime.now()-datetime.fromisoformat(dt_str)).total_seconds()/3600)
        except: return 0

    @staticmethod
    def _hours_between(s, e):
        if not s or not e: return None
        try:
            return max(0,(datetime.fromisoformat(e)-datetime.fromisoformat(s)).total_seconds()/3600)
        except: return None

    @staticmethod
    def _age_colour(hours):
        if hours < 4:  return "#1A6B3A", "#F0FFF4"   # green
        if hours < 24: return "#B45309", "#FFF8E7"   # amber
        return          "#B71C1C", "#FFF0F0"          # red

    @staticmethod
    def _fmt_hrs(h):
        if h is None: return "—"
        if h < 1:     return "< 1h"
        if h < 24:    return f"{h:.1f}h"
        return f"{int(h//24)}d {int(h%24)}h"

    def _build(self):
        tk.Frame(self, bg=FBC_ACCENT, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="📊  Aging & Dealer Performance Analysis",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,12,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text="Green < 4h  |  Amber 4–24h  |  Red > 24h",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))

        # ── Build data ────────────────────────────────────────────────
        perf = {}
        for o in self.all_orders:
            dealer = o.get("taken_by","") or ""
            if not dealer: continue
            if dealer not in perf:
                perf[dealer] = {"taken":0,"executed":0,"partial":0,
                                "cancelled":0,"exec_hours":[],
                                "pending_hours":[]}
            perf[dealer]["taken"] += 1
            s = o.get("status","")
            # Time from TAKEN to EXECUTED
            h = self._hours_between(o.get("taken_datetime",""),
                                    o.get("executed_datetime",""))
            if s == "EXECUTED":
                perf[dealer]["executed"] += 1
                if h is not None: perf[dealer]["exec_hours"].append(h)
            elif s == "PARTIAL":
                perf[dealer]["partial"] += 1
            elif s == "CANCELLED":
                perf[dealer]["cancelled"] += 1

        open_orders = sorted(
            [o for o in self.all_orders if o.get("status") in ("PENDING","TAKEN","PARTIAL")],
            key=lambda o: o.get("entered_datetime","")
        )

        # ── Layout: left panel + right canvas ─────────────────────────
        main = tk.Frame(self, bg=BG); main.pack(fill="both", expand=True)

        left = tk.Frame(main, bg=BG, width=640)
        left.pack(side="left", fill="both", expand=True)
        left.pack_propagate(False)

        right_outer = tk.Frame(main, bg=BG, width=440)
        right_outer.pack(side="right", fill="both")
        right_outer.pack_propagate(False)
        _, right, _ = scrollable_body(right_outer)
        _, body, _ = scrollable_body(left)

        # ── DEALER PERFORMANCE TABLE ──────────────────────────────────
        tk.Label(body, text="DEALER PERFORMANCE", bg=BG, fg=FBC_MID,
                 font=(FONT,9,"bold")).pack(anchor="w", pady=(6,4))

        if perf:
            tbl = tk.Frame(body, bg=CARD_BG, highlightbackground=SEP_CLR, highlightthickness=1)
            tbl.pack(fill="x", pady=(0,12))
            hrow = tk.Frame(tbl, bg=TBL_HDR_BG); hrow.pack(fill="x")
            for txt, w in [("Dealer",16),("Taken",7),("Executed",9),
                           ("Partial",8),("Cancelled",9),("Avg Speed",13),
                           ("Fastest",11),("Slowest",11)]:
                tk.Label(hrow, text=txt, bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         font=(FONT,8,"bold"), width=w, anchor="w",
                         padx=4, pady=5).pack(side="left")
            for i, (dealer, d) in enumerate(sorted(perf.items())):
                rbg = TBL_ROW_A if i%2==0 else TBL_ROW_B
                hrs = d["exec_hours"]
                avg_h  = sum(hrs)/len(hrs) if hrs else None
                fast_h = min(hrs) if hrs else None
                slow_h = max(hrs) if hrs else None
                is_me  = (dealer == self.dealer_name)
                rrow = tk.Frame(tbl, bg=rbg); rrow.pack(fill="x")
                name_txt = f"★ {dealer}" if is_me else dealer
                speed_fg = "#B71C1C" if avg_h and avg_h > 24 else \
                           "#B45309" if avg_h and avg_h > 4 else "#1A6B3A"
                for txt, w, fg in [
                    (name_txt, 16, FBC_MID if is_me else "#1A2B3C"),
                    (str(d["taken"]),    7,  "#1A2B3C"),
                    (str(d["executed"]), 9,  "#1A6B3A"),
                    (str(d["partial"]),  8,  "#6B21A8"),
                    (str(d["cancelled"]),9,  "#B71C1C"),
                    (self._fmt_hrs(avg_h),  13, speed_fg),
                    (self._fmt_hrs(fast_h), 11, "#1A6B3A"),
                    (self._fmt_hrs(slow_h), 11,
                     "#B71C1C" if slow_h and slow_h>24 else "#607080"),
                ]:
                    tk.Label(rrow, text=txt, bg=rbg, fg=fg,
                             font=(FONT,9,"bold" if is_me else "normal"),
                             width=w, anchor="w", padx=4, pady=5).pack(side="left")
        else:
            tk.Label(body, text="No dealer activity recorded yet.",
                     bg=BG, fg="#8096B0", font=(FONT,9)).pack(anchor="w")

        # ── OPEN ORDERS — OLDEST FIRST ────────────────────────────────
        tk.Label(body, text="OPEN ORDERS — OLDEST FIRST", bg=BG, fg=FBC_MID,
                 font=(FONT,9,"bold")).pack(anchor="w", pady=(14,4))

        if not open_orders:
            tk.Label(body, text="No open orders — great work!",
                     bg=BG, fg="#1A6B3A", font=(FONT,9)).pack(anchor="w")
        else:
            ohdr = tk.Frame(body, bg=TBL_HDR_BG); ohdr.pack(fill="x")
            for txt, w in [("Total Age",11),("Status",13),("Type",5),("Counter",10),
                           ("Client",17),("Shares",10),("Entered By",11),
                           ("Dealer",13),("Stage",12)]:
                tk.Label(ohdr, text=txt, bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                         font=(FONT,8,"bold"), width=w, anchor="w",
                         padx=4, pady=5).pack(side="left")

            for i, o in enumerate(open_orders):
                rbg = TBL_ROW_A if i%2==0 else TBL_ROW_B
                status = o.get("status","")
                total_hrs = self._hours_since(o.get("entered_datetime",""))
                age_fg, age_bg = self._age_colour(total_hrs)

                if status == "PENDING":
                    stage_hrs = self._hours_since(o.get("entered_datetime",""))
                    stage_lbl = "Pending"
                else:
                    stage_hrs = self._hours_since(
                        o.get("taken_datetime","") or o.get("entered_datetime",""))
                    stage_lbl = "In Progress" if status=="TAKEN" else "Partial"
                stg_fg, stg_bg = self._age_colour(stage_hrs)

                status_cfg = {
                    "PENDING": ("PENDING",     "#B45309","#FEF3C7"),
                    "TAKEN":   ("IN PROGRESS", "#0066B3","#DBEAFE"),
                    "PARTIAL": ("PARTIAL",     "#6B21A8","#F3E8FF"),
                }
                st_txt, st_fg, st_bg = status_cfg.get(status,(status,"#607080",BG))

                orow = tk.Frame(body, bg=rbg); orow.pack(fill="x")
                tk.Frame(orow, bg=age_fg, width=4).pack(side="left", fill="y")

                # Total age pill
                tk.Label(orow, text=self._fmt_hrs(total_hrs),
                         bg=age_bg, fg=age_fg, font=(FONT,8,"bold"),
                         width=11, anchor="w", padx=4, pady=5).pack(side="left")
                # Status pill
                tk.Label(orow, text=st_txt, bg=st_bg, fg=st_fg,
                         font=(FONT,8,"bold"), width=13, anchor="w",
                         padx=4, pady=5).pack(side="left")

                def cell(txt, w, fg="#1A2B3C", bg=rbg):
                    tk.Label(orow, text=str(txt) if txt else "—", bg=bg, fg=fg,
                             font=(FONT,9), width=w, anchor="w",
                             padx=4, pady=5).pack(side="left")

                ot = o.get("order_type","")
                ot_bg = "#1A6B3A" if ot=="BUY" else "#B71C1C"
                tk.Label(orow, text=ot, bg=ot_bg, fg=WHITE,
                         font=(FONT,8,"bold"), width=5, anchor="center",
                         pady=5).pack(side="left")

                client = o.get("client_name","")
                if len(client) > 16: client = client[:15]+"…"
                cell(o.get("counter",""), 10, fg=FBC_DARK)
                cell(client, 17)
                cell(f"{o.get('num_shares',0):,}", 10)
                cell(o.get("entered_by",""), 11, fg="#607080")

                taken = o.get("taken_by","")
                cell(taken or "—", 13, fg="#C0392B" if taken else "#C0C8D8")

                tk.Label(orow, text=f"{stage_lbl}: {self._fmt_hrs(stage_hrs)}",
                         bg=stg_bg, fg=stg_fg, font=(FONT,8),
                         width=12, anchor="w", padx=4, pady=5).pack(side="left")

                tk.Frame(body, bg=SEP_CLR, height=1).pack(fill="x")

        tk.Frame(body, bg=BG, height=10).pack()

        # Close button outside scroll area at bottom
        btn_bar = tk.Frame(self, bg=BG, pady=8)
        btn_bar.pack(fill="x", side="bottom")
        tk.Button(btn_bar, text="  Close  ", font=(FONT, 10), bg=FBC_MID, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_DARK,
                  command=self.destroy, pady=5).pack()

        # ── RIGHT PANEL: CHARTS ───────────────────────────────────────
        tk.Label(right, text="VISUAL ANALYTICS", bg=BG, fg=FBC_MID,
                 font=(FONT,9,"bold")).pack(anchor="w", padx=12, pady=(10,6))

        self._draw_charts(right, perf, open_orders)

    def _draw_charts(self, parent, perf, open_orders):
        """Draw bar charts using pure tkinter Canvas."""

        def bar_chart(container, title, labels, values, colours, max_val=None, unit=""):
            frame = tk.Frame(container, bg=CARD_BG,
                             highlightbackground=SEP_CLR, highlightthickness=1)
            frame.pack(fill="x", padx=12, pady=(0,10))
            tk.Label(frame, text=title, bg=CARD_BG, fg=FBC_DARK,
                     font=(FONT,8,"bold")).pack(anchor="w", padx=8, pady=(6,2))

            if not values or max(values or [0]) == 0:
                tk.Label(frame, text="No data yet", bg=CARD_BG,
                         fg="#8096B0", font=(FONT,8), pady=8).pack()
                return

            cw, ch = 370, 110
            mv = max_val or max(values)
            canvas = tk.Canvas(frame, width=cw, height=ch, bg=CARD_BG,
                               highlightthickness=0)
            canvas.pack(padx=8, pady=(0,8))

            pad_l, pad_r, pad_t, pad_b = 50, 10, 10, 30
            chart_w = cw - pad_l - pad_r
            chart_h = ch - pad_t - pad_b
            n = len(labels)
            bar_w = max(8, (chart_w // n) - 6)

            # Grid lines
            for step in [0.25, 0.5, 0.75, 1.0]:
                y = pad_t + chart_h - int(chart_h * step)
                canvas.create_line(pad_l, y, cw-pad_r, y,
                                   fill="#E0E8F0", dash=(2,4))
                val = mv * step
                label = f"{val:.0f}{unit}" if val == int(val) else f"{val:.1f}{unit}"
                canvas.create_text(pad_l-4, y, text=label,
                                   anchor="e", font=(FONT,7), fill="#8096B0")

            # Bars
            for i, (lbl, val, col) in enumerate(zip(labels, values, colours)):
                x0 = pad_l + i*(chart_w//n) + (chart_w//n - bar_w)//2
                x1 = x0 + bar_w
                bar_h = int(chart_h * min(val/mv, 1.0)) if mv else 0
                y0 = pad_t + chart_h - bar_h
                y1 = pad_t + chart_h
                if bar_h > 0:
                    canvas.create_rectangle(x0, y0, x1, y1, fill=col, outline="")
                    # Value label on top of bar
                    vstr = f"{val:.1f}{unit}" if isinstance(val,float) else str(val)
                    canvas.create_text((x0+x1)//2, y0-3, text=vstr,
                                       anchor="s", font=(FONT,7,"bold"), fill=col)
                # X label
                short_lbl = lbl[:8] if len(lbl)>8 else lbl
                canvas.create_text((x0+x1)//2, pad_t+chart_h+4,
                                   text=short_lbl, anchor="n",
                                   font=(FONT,7), fill="#607080")

        # Chart 1: Orders taken per dealer
        dealers = sorted(perf.keys())
        if dealers:
            bar_chart(parent,
                      "Orders Taken per Dealer",
                      dealers,
                      [perf[d]["taken"] for d in dealers],
                      [FBC_MID]*len(dealers))

        # Chart 2: Avg execution speed per dealer (hours)
        spd_dealers = [d for d in dealers if perf[d]["exec_hours"]]
        if spd_dealers:
            avgs = [sum(perf[d]["exec_hours"])/len(perf[d]["exec_hours"])
                    for d in spd_dealers]
            cols = ["#1A6B3A" if a < 4 else "#B45309" if a < 24 else "#B71C1C"
                    for a in avgs]
            bar_chart(parent,
                      "Avg Execution Speed per Dealer",
                      spd_dealers, avgs, cols, unit="h")

        # Chart 3: Open order age distribution
        if open_orders:
            buckets = {"< 4h": 0, "4–24h": 0, "1–2d": 0, "> 2d": 0}
            for o in open_orders:
                h = self._hours_since(o.get("entered_datetime",""))
                if   h < 4:   buckets["< 4h"]  += 1
                elif h < 24:  buckets["4–24h"] += 1
                elif h < 48:  buckets["1–2d"]  += 1
                else:         buckets["> 2d"]  += 1
            cols3 = ["#1A6B3A","#B45309","#E67E22","#B71C1C"]
            bar_chart(parent,
                      "Open Order Age Distribution",
                      list(buckets.keys()),
                      list(buckets.values()),
                      cols3)

        # Chart 4: Executed vs Partial vs Cancelled per dealer
        if dealers:
            # Stacked-style: show executed count coloured by rate
            exec_counts = [perf[d]["executed"] for d in dealers]
            exec_rate_cols = []
            for d in dealers:
                t = perf[d]["taken"]
                r = perf[d]["executed"]/t if t else 0
                exec_rate_cols.append(
                    "#1A6B3A" if r >= 0.8 else "#B45309" if r >= 0.5 else "#B71C1C")
            bar_chart(parent,
                      "Executed Orders per Dealer",
                      dealers, exec_counts, exec_rate_cols)

        # Summary stats box
        summary = tk.Frame(parent, bg=CARD_BG,
                           highlightbackground=SEP_CLR, highlightthickness=1)
        summary.pack(fill="x", padx=12, pady=(0,10))
        tk.Label(summary, text="SUMMARY", bg=CARD_BG, fg=FBC_DARK,
                 font=(FONT,8,"bold")).pack(anchor="w", padx=8, pady=(6,4))

        total_open   = len(open_orders)
        overdue      = sum(1 for o in open_orders
                          if self._hours_since(o.get("entered_datetime","")) >= 48)
        all_hours    = [h for d in perf.values() for h in d["exec_hours"]]
        global_avg   = sum(all_hours)/len(all_hours) if all_hours else None
        fastest_exec = min(all_hours) if all_hours else None
        slowest_exec = max(all_hours) if all_hours else None

        for lbl, val, fg in [
            ("Open orders:",       str(total_open),          FBC_DARK),
            ("Overdue (> 2d):",    str(overdue),
             "#B71C1C" if overdue else "#1A6B3A"),
            ("Global avg speed:",  self._fmt_hrs(global_avg), "#B45309"),
            ("Fastest execution:", self._fmt_hrs(fastest_exec),"#1A6B3A"),
            ("Slowest execution:", self._fmt_hrs(slowest_exec),
             "#B71C1C" if slowest_exec and slowest_exec>24 else "#607080"),
        ]:
            r = tk.Frame(summary, bg=CARD_BG); r.pack(fill="x", padx=8, pady=1)
            tk.Label(r, text=lbl, bg=CARD_BG, fg="#607080",
                     font=(FONT,8), width=18, anchor="w").pack(side="left")
            tk.Label(r, text=val, bg=CARD_BG, fg=fg,
                     font=(FONT,9,"bold")).pack(side="left")
        tk.Frame(summary, bg=BG, height=6).pack()
# ════════════════════════════════════════════════════════════════════════════
#  ORDER SLIP — PDF GENERATION
# ════════════════════════════════════════════════════════════════════════════
class OrderSlipPDF:
    """Generates a printable PDF order slip."""

    @staticmethod
    def generate(order, save_path):
        try:
            from reportlab.lib.pagesizes import A5
            from reportlab.lib import colors
            from reportlab.platypus import (SimpleDocTemplate, Table,
                                            TableStyle, Paragraph,
                                            Spacer, HRFlowable)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
        except ImportError:
            return False, "reportlab not installed. Run: pip install reportlab"

        try:
            doc = SimpleDocTemplate(save_path, pagesize=A5,
                                    leftMargin=12*mm, rightMargin=12*mm,
                                    topMargin=10*mm, bottomMargin=10*mm)
            styles = getSampleStyleSheet()
            story  = []

            # ── Header ──────────────────────────────────────────────
            hdr_style = ParagraphStyle("hdr", fontSize=13, fontName="Helvetica-Bold",
                                       textColor=colors.HexColor("#003B6F"),
                                       spaceAfter=2)
            sub_style = ParagraphStyle("sub", fontSize=8, fontName="Helvetica",
                                       textColor=colors.HexColor("#607080"),
                                       spaceAfter=6)
            val_style = ParagraphStyle("val", fontSize=9, fontName="Helvetica",
                                       textColor=colors.black)
            lbl_style = ParagraphStyle("lbl", fontSize=8, fontName="Helvetica",
                                       textColor=colors.HexColor("#607080"))

            story.append(Paragraph("FBC Securities", hdr_style))
            story.append(Paragraph("Order Management System — Official Order Slip", sub_style))
            story.append(HRFlowable(width="100%", thickness=2,
                                    color=colors.HexColor("#003B6F")))
            story.append(Spacer(1, 4*mm))

            ot  = order.get("order_type", "")
            ot_colour = colors.HexColor("#1A6B3A") if ot == "BUY" \
                        else colors.HexColor("#B71C1C")

            # Order type badge
            badge_style = ParagraphStyle("badge", fontSize=14,
                                         fontName="Helvetica-Bold",
                                         textColor=ot_colour, spaceAfter=4)
            story.append(Paragraph(f"◼  {ot} ORDER", badge_style))

            # ── Order details table ──────────────────────────────────
            def row(label, value):
                return [Paragraph(label, lbl_style),
                        Paragraph(str(value) if value else "—", val_style)]

            mode = order.get("amount_mode", "SHARES")
            qty_str = (f"{order.get('total_amount','')} {order.get('exchange','')}"
                       if mode == "AMOUNT"
                       else f"{order.get('num_shares', 0):,} shares")

            order_data = [
                row("Order ID",       order.get("id", "")),
                row("Counter",        order.get("counter", "")),
                row("Exchange",       order.get("exchange", "")),
                row("Quantity",       qty_str),
                row("Limit Price",    order.get("limit_price", "")),
                row("Order Date",     fmt_date(order.get("order_date", ""))),
                row("Entered By",     order.get("entered_by", "")),
                row("Entered At",     fmt_dt(order.get("entered_datetime", ""))),
                row("Status",         order.get("status", "")),
            ]
            if order.get("shares_executed", 0) > 0:
                order_data.append(
                    row("Shares Filled",
                        f"{order['shares_executed']:,} @ {order.get('execution_price','')}"))
            if order.get("taken_by"):
                order_data.append(row("Dealer",    order.get("taken_by", "")))
                order_data.append(row("Taken At",  fmt_dt(order.get("taken_datetime", ""))))
            if order.get("executed_by"):
                order_data.append(row("Executed By", order.get("executed_by", "")))
                order_data.append(row("Executed At", fmt_dt(order.get("executed_datetime", ""))))

            tbl = Table(order_data, colWidths=[35*mm, 75*mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#EEF2F7")),
                ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS", (0,0), (-1,-1),
                 [colors.white, colors.HexColor("#F8FAFC")]),
                ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#D0DAE8")),
                ("LEFTPADDING",  (0,0), (-1,-1), 4),
                ("RIGHTPADDING", (0,0), (-1,-1), 4),
                ("TOPPADDING",   (0,0), (-1,-1), 3),
                ("BOTTOMPADDING",(0,0), (-1,-1), 3),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 4*mm))

            # ── Client details ───────────────────────────────────────
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=colors.HexColor("#D0DAE8")))
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph("CLIENT DETAILS", ParagraphStyle(
                "sec", fontSize=8, fontName="Helvetica-Bold",
                textColor=colors.HexColor("#0066B3"), spaceAfter=3)))

            client_data = [
                row("Client Name",    order.get("client_name", "")),
                row("Client Address", order.get("client_address", "")),
                row("CSD Number",     order.get("csd_no", "")),
                row("Custodian",      order.get("custodian", "")),
                row("Instruction By", order.get("instruction_by", "")),
            ]
            if order.get("tel_no"):
                client_data.append(row("Tel No.", order.get("tel_no", "")))

            ctbl = Table(client_data, colWidths=[35*mm, 75*mm])
            ctbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#EEF2F7")),
                ("FONTNAME",   (0,0), (-1,-1), "Helvetica"),
                ("FONTSIZE",   (0,0), (-1,-1), 8),
                ("ROWBACKGROUNDS", (0,0), (-1,-1),
                 [colors.white, colors.HexColor("#F8FAFC")]),
                ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#D0DAE8")),
                ("LEFTPADDING",  (0,0), (-1,-1), 4),
                ("RIGHTPADDING", (0,0), (-1,-1), 4),
                ("TOPPADDING",   (0,0), (-1,-1), 3),
                ("BOTTOMPADDING",(0,0), (-1,-1), 3),
            ]))
            story.append(ctbl)

            # ── Notes ───────────────────────────────────────────────
            if order.get("notes"):
                story.append(Spacer(1, 3*mm))
                story.append(Paragraph("NOTES", ParagraphStyle(
                    "sec2", fontSize=8, fontName="Helvetica-Bold",
                    textColor=colors.HexColor("#0066B3"), spaceAfter=2)))
                story.append(Paragraph(order["notes"], ParagraphStyle(
                    "notes", fontSize=8, fontName="Helvetica",
                    textColor=colors.HexColor("#1A2B3C"))))

            # ── Signature lines ──────────────────────────────────────
            story.append(Spacer(1, 8*mm))
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=colors.HexColor("#D0DAE8")))
            story.append(Spacer(1, 4*mm))

            sig_data = [[
                Paragraph("Dealer Signature: ___________________",
                          ParagraphStyle("sig", fontSize=8, fontName="Helvetica",
                                         textColor=colors.HexColor("#607080"))),
                Paragraph("Date: _______________",
                          ParagraphStyle("sig2", fontSize=8, fontName="Helvetica",
                                         textColor=colors.HexColor("#607080"))),
            ]]
            sig_tbl = Table(sig_data, colWidths=[75*mm, 35*mm])
            story.append(sig_tbl)
            story.append(Spacer(1, 3*mm))
            story.append(Paragraph(
                f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}  |  "
                f"Order Manager v{VERSION}  |  FBC Securities",
                ParagraphStyle("footer", fontSize=6, fontName="Helvetica",
                               textColor=colors.HexColor("#A0B0C0"),
                               alignment=1)))

            doc.build(story)
            return True, ""
        except Exception as e:
            return False, str(e)
# ════════════════════════════════════════════════════════════════════════════
#  DAILY REPORT EXPORT
# ════════════════════════════════════════════════════════════════════════════
class DailyReportExporter:
    """Exports daily order report to Excel/CSV."""

    @staticmethod
    def export(orders, report_date, save_path):
        try:
            import csv as _csv_mod

            # Filter to report date
            date_str = report_date.isoformat()
            day_orders = [o for o in orders
                          if o.get("order_date","") == date_str
                          or o.get("entered_datetime","")[:10] == date_str]

            # Also include orders executed on this date
            exec_orders = [o for o in orders
                           if o.get("executed_datetime","")[:10] == date_str
                           and o not in day_orders]
            all_relevant = day_orders + exec_orders

            if not all_relevant:
                return False, f"No orders found for {report_date.strftime('%d %b %Y')}"

            # Try Excel first, fall back to CSV
            try:
                import openpyxl
                from openpyxl.styles import (Font, PatternFill,
                                              Alignment, Border, Side)
                from openpyxl.utils import get_column_letter
                DailyReportExporter._export_excel(
                    all_relevant, orders, report_date, save_path, day_orders)
                return True, ""
            except ImportError:
                # Fall back to CSV
                csv_path = save_path.replace(".xlsx", ".csv")
                DailyReportExporter._export_csv(all_relevant, csv_path)
                return True, f"(openpyxl not installed — saved as CSV instead:\n{csv_path})"

        except Exception as e:
            return False, str(e)

    @staticmethod
    def _export_csv(orders, path):
        import csv as _csv_mod
        fields = ["id","order_type","client_name","csd_no","counter",
                  "exchange","num_shares","limit_price","status",
                  "entered_by","taken_by","executed_by",
                  "shares_executed","execution_price",
                  "order_date","entered_datetime","executed_datetime"]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = _csv_mod.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(orders)

    @staticmethod
    def _export_excel(day_orders, all_orders, report_date, path, entered_today):
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side)
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Colour helpers ───────────────────────────────────────────
        def fill(hex_col):
            return PatternFill("solid", fgColor=hex_col.lstrip("#"))

        def hdr_cell(ws, row, col, text, bg="002855", fg="90CAF9", bold=True, width=None):
            c = ws.cell(row=row, column=col, value=text)
            c.font = Font(bold=bold, color=fg, size=9)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            thin = Side(style="thin", color="D0DAE8")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if width:
                ws.column_dimensions[get_column_letter(col)].width = width
            return c

        def data_cell(ws, row, col, value, fg="1A2B3C",
                      bg=None, bold=False, align="left"):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(color=fg, size=9, bold=bold)
            if bg: c.fill = fill(bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            thin = Side(style="thin", color="D0DAE8")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            return c

        STATUS_COLOURS = {
            "PENDING":   ("B45309","FEF3C7"),
            "TAKEN":     ("0066B3","DBEAFE"),
            "EXECUTED":  ("1A6B3A","DCFCE7"),
            "PARTIAL":   ("6B21A8","F3E8FF"),
            "CANCELLED": ("757575","F5F5F5"),
        }

        date_label = report_date.strftime("%d %B %Y")

        # ══ Sheet 1: Orders Summary ══════════════════════════════════
        ws1 = wb.active; ws1.title = "Orders Summary"
        ws1.row_dimensions[1].height = 30
        ws1.row_dimensions[2].height = 20

        # Title
        ws1.merge_cells("A1:L1")
        title = ws1["A1"]
        title.value = f"FBC Securities — Daily Order Report    {date_label}"
        title.font  = Font(bold=True, size=13, color="003B6F")
        title.alignment = Alignment(horizontal="left", vertical="center")
        title.fill = fill("EEF2F7")

        ws1.merge_cells("A2:L2")
        sub = ws1["A2"]
        sub.value = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}  |  FBC Order Manager v{VERSION}"
        sub.font  = Font(size=8, color="607080")
        sub.alignment = Alignment(horizontal="left", vertical="center")

        # Column headers row 3
        cols = [
            ("Order ID",      8),  ("Type",    6),  ("Counter",   10),
            ("Client Name",  22),  ("CSD No.", 18),  ("Exchange",   8),
            ("Shares",       10),  ("Limit",    9),  ("Status",    12),
            ("Entered By",   12),  ("Dealer",  12),  ("Exec Price", 10),
        ]
        for ci, (label, width) in enumerate(cols, 1):
            hdr_cell(ws1, 3, ci, label, width=width)

        ws1.freeze_panes = "A4"

        for ri, o in enumerate(day_orders, 4):
            ws1.row_dimensions[ri].height = 16
            status = o.get("status","")
            fg_s, bg_s = STATUS_COLOURS.get(status, ("1A2B3C","FFFFFF"))
            ot = o.get("order_type","")
            ot_fg = "1A6B3A" if ot == "BUY" else "B71C1C"
            row_bg = "FFFFFF" if ri % 2 == 0 else "F4F7FB"

            data_cell(ws1, ri, 1,  o.get("id",""),          bold=True, align="center")
            data_cell(ws1, ri, 2,  ot,     fg=ot_fg, bold=True, align="center")
            data_cell(ws1, ri, 3,  o.get("counter",""),     bold=True)
            data_cell(ws1, ri, 4,  o.get("client_name",""))
            data_cell(ws1, ri, 5,  o.get("csd_no",""),      fg="607080")
            data_cell(ws1, ri, 6,  o.get("exchange",""),    align="center")
            data_cell(ws1, ri, 7,  o.get("num_shares",0),   align="right")
            data_cell(ws1, ri, 8,  o.get("limit_price",""), align="right")
            sc = data_cell(ws1, ri, 9, status, fg=fg_s, bg=bg_s,
                           bold=True, align="center")
            data_cell(ws1, ri, 10, o.get("entered_by",""),  fg="607080")
            data_cell(ws1, ri, 11, o.get("taken_by","") or o.get("executed_by",""),
                      fg="C0392B" if o.get("taken_by") else "607080")
            data_cell(ws1, ri, 12,
                      f"{o.get('shares_executed',0):,} @ {o.get('execution_price','')}"
                      if o.get("shares_executed",0) > 0 else "—",
                      fg="1A6B3A", align="right")

        # ══ Sheet 2: Dealer Performance ══════════════════════════════
        ws2 = wb.create_sheet("Dealer Performance")
        ws2.merge_cells("A1:H1")
        t2 = ws2["A1"]
        t2.value = f"Dealer Performance — {date_label}"
        t2.font  = Font(bold=True, size=12, color="003B6F")
        t2.fill  = fill("EEF2F7")
        t2.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[1].height = 25

        d_cols = [("Dealer",16),("Orders Taken",14),("Executed",10),
                  ("Partial",10),("Cancelled",10),("Exec Rate",10),
                  ("Total Shares Exec",16),("Notes",20)]
        for ci, (lbl, w) in enumerate(d_cols, 1):
            hdr_cell(ws2, 2, ci, lbl, width=w)
        ws2.freeze_panes = "A3"

        # Build dealer stats from all_orders for the day
        perf = {}
        for o in all_orders:
            dealer = o.get("taken_by","") or ""
            if not dealer: continue
            od = o.get("order_date","") or o.get("entered_datetime","")[:10]
            if od != report_date.isoformat(): continue
            if dealer not in perf:
                perf[dealer] = {"taken":0,"executed":0,"partial":0,
                                "cancelled":0,"shares_exec":0}
            perf[dealer]["taken"] += 1
            s = o.get("status","")
            if s == "EXECUTED":
                perf[dealer]["executed"]   += 1
                perf[dealer]["shares_exec"] += o.get("shares_executed",0)
            elif s == "PARTIAL":
                perf[dealer]["partial"]    += 1
                perf[dealer]["shares_exec"] += o.get("shares_executed",0)
            elif s == "CANCELLED":
                perf[dealer]["cancelled"]  += 1

        for ri2, (dealer, d) in enumerate(sorted(perf.items()), 3):
            rate = d["executed"] / d["taken"] if d["taken"] else 0
            rate_fg = "1A6B3A" if rate >= 0.8 else \
                      "B45309" if rate >= 0.5 else "B71C1C"
            note = ""
            if rate < 0.5: note = "Low execution rate — review"
            rb = "FFFFFF" if ri2 % 2 == 0 else "F4F7FB"
            data_cell(ws2, ri2, 1, dealer,             bold=True)
            data_cell(ws2, ri2, 2, d["taken"],         align="center")
            data_cell(ws2, ri2, 3, d["executed"],      fg="1A6B3A", align="center")
            data_cell(ws2, ri2, 4, d["partial"],       fg="6B21A8", align="center")
            data_cell(ws2, ri2, 5, d["cancelled"],     fg="B71C1C", align="center")
            data_cell(ws2, ri2, 6, f"{rate*100:.0f}%", fg=rate_fg,
                      bold=True, align="center")
            data_cell(ws2, ri2, 7, f"{d['shares_exec']:,}", align="right")
            data_cell(ws2, ri2, 8, note, fg="B45309")

        # ══ Sheet 3: Summary Stats ════════════════════════════════════
        ws3 = wb.create_sheet("Summary")
        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 20

        ws3.merge_cells("A1:B1")
        t3 = ws3["A1"]
        t3.value = f"End-of-Day Summary — {date_label}"
        t3.font  = Font(bold=True, size=12, color="003B6F")
        t3.fill  = fill("EEF2F7")
        t3.alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[1].height = 25

        stats = [
            ("Total orders entered today",
             len(entered_today)),
            ("BUY orders",
             len([o for o in day_orders if o.get("order_type")=="BUY"])),
            ("SELL orders",
             len([o for o in day_orders if o.get("order_type")=="SELL"])),
            ("Orders executed",
             len([o for o in day_orders if o.get("status")=="EXECUTED"])),
            ("Orders partially filled",
             len([o for o in day_orders if o.get("status")=="PARTIAL"])),
            ("Orders still pending",
             len([o for o in day_orders if o.get("status")=="PENDING"])),
            ("Orders cancelled",
             len([o for o in day_orders if o.get("status")=="CANCELLED"])),
            ("ZSE orders",
             len([o for o in day_orders if o.get("exchange")=="ZSE"])),
            ("VFEX orders",
             len([o for o in day_orders if o.get("exchange")=="VFEX"])),
            ("Total shares ordered (ZSE)",
             sum(o.get("num_shares",0) for o in day_orders
                 if o.get("exchange")=="ZSE")),
            ("Total shares ordered (VFEX)",
             sum(o.get("num_shares",0) for o in day_orders
                 if o.get("exchange")=="VFEX")),
            ("Total shares executed",
             sum(o.get("shares_executed",0) for o in day_orders)),
        ]

        thin = Side(style="thin", color="D0DAE8")
        for ri3, (label, value) in enumerate(stats, 2):
            ws3.row_dimensions[ri3].height = 16
            lc = ws3.cell(row=ri3, column=1, value=label)
            lc.font = Font(size=9, color="607080")
            lc.fill = fill("EEF2F7")
            lc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            vc = ws3.cell(row=ri3, column=2, value=value)
            vc.font = Font(size=9, bold=True, color="003B6F")
            vc.alignment = Alignment(horizontal="right")
            vc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if isinstance(value, int):
                vc.number_format = "#,##0"

        wb.save(path)
class DailyReportDialog(tk.Toplevel):
    def __init__(self, parent, orders):
        super().__init__(parent)
        self.orders = orders
        self.title("Export Daily Report")
        self.resizable(False, False); self.configure(bg=BG); self.grab_set()
        self._build()
        self.update_idletasks()
        w, h = 460, 300
        px = parent.winfo_rootx() + (parent.winfo_width()  - w) // 2
        py = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    def _build(self):
        tk.Frame(self, bg=FBC_ACCENT, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="📊  Export Daily Report",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,12,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text="Exports to Excel (.xlsx) — 3 sheets: Orders, Dealer Performance, Summary",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))

        body = tk.Frame(self, bg=BG, padx=24, pady=16); body.pack(fill="both", expand=True)

        info = tk.Frame(body, bg="#E8F1FB", padx=12, pady=10,
                        highlightbackground="#A8C4E0", highlightthickness=1)
        info.pack(fill="x", pady=(0,16))
        tk.Label(info,
                 text="Select the date to report on.\n"
                      "Includes all orders entered OR executed on that date.\n"
                      "Install openpyxl for Excel: pip install openpyxl",
                 bg="#E8F1FB", fg=FBC_DARK, font=(FONT,9), justify="left").pack(anchor="w")

        r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=4)
        tk.Label(r, text="Report date:", bg=BG, fg="#607080",
                 font=(FONT,9), width=14, anchor="w").pack(side="left")
        self._date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        date_entry(r, self._date_var, bg=BG).pack(side="left", padx=(4,0))

        self._err = tk.Label(body, text="", bg=BG, fg="#B71C1C", font=(FONT,9))
        self._err.pack(anchor="w", pady=(8,0))

        btn_bar = tk.Frame(self, bg=BG, padx=20, pady=10); btn_bar.pack(fill="x")
        tk.Button(btn_bar, text="Cancel", font=(FONT,10), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", command=self.destroy,
                  activebackground=SEP_CLR).pack(side="right", padx=(6,0))
        tk.Button(btn_bar, text="  📊  Export Report  ", font=(FONT,10,"bold"),
                  bg=FBC_MID, fg=WHITE, relief="flat", cursor="hand2",
                  activebackground=FBC_DARK, command=self._export).pack(side="right")

    def _export(self):
        d = parse_date_input(self._date_var.get())
        if not d:
            self._err.config(text="❌  Please enter a valid date (dd/mm/yyyy)."); return
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        save_dir  = downloads if os.path.isdir(downloads) else os.path.expanduser("~")
        filename  = f"FBC_DailyReport_{d.strftime('%Y%m%d')}.xlsx"
        save_path = os.path.join(save_dir, filename)
        self._err.config(text="⏳  Generating report..."); self.update()
        ok, msg = DailyReportExporter.export(self.orders, d, save_path)
        if ok:
            info = f"Report saved to:\n{save_path}"
            if msg: info += f"\n\n{msg}"
            messagebox.showinfo("Report Exported", info, parent=self)
            try: os.startfile(save_path)
            except: pass
            self.destroy()
        else:
            self._err.config(text=f"❌  {msg[:80]}")
# ════════════════════════════════════════════════════════════════════════════
#  CLIENT ORDER HISTORY DIALOG
# ════════════════════════════════════════════════════════════════════════════
class ClientHistoryDialog(tk.Toplevel):
    def __init__(self, parent, client_name, csd_no, all_orders):
        super().__init__(parent)
        self.title(f"Order History — {client_name}")
        self.resizable(True, True); self.configure(bg=BG); self.grab_set()
        self.client_name = client_name
        self.csd_no      = csd_no
        self.all_orders  = all_orders
        self._build()
        self.update_idletasks()
        w, h = 900, 580
        px = parent.winfo_rootx() + (parent.winfo_width()  - w) // 2
        py = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{max(px,0)}+{max(py,0)}")

    def _build(self):
        tk.Frame(self, bg=FBC_ACCENT, height=4).pack(fill="x")
        hdr = tk.Frame(self, bg=FBC_DARK, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"📋  Order History — {self.client_name}",
                 bg=FBC_DARK, fg=WHITE, font=(FONT,12,"bold")).pack(padx=16, anchor="w")
        tk.Label(hdr, text=f"CSD: {self.csd_no}   ·   All orders ever placed for this client",
                 bg=FBC_DARK, fg=SIDEBAR_TXT, font=(FONT,9)).pack(padx=16, anchor="w", pady=(0,4))

        # Filter to this client
        client_orders = sorted(
            [o for o in self.all_orders
             if o.get("csd_no","").strip().upper() == self.csd_no.strip().upper()
             or o.get("client_name","").strip().upper() == self.client_name.strip().upper()],
            key=lambda o: o.get("entered_datetime",""),
            reverse=True
        )

        # ── Summary cards ────────────────────────────────────────────
        stats_bar = tk.Frame(self, bg=SIDEBAR_BG, pady=6)
        stats_bar.pack(fill="x")
        sr = tk.Frame(stats_bar, bg=SIDEBAR_BG); sr.pack(padx=16)

        def scard(label, value, colour):
            f = tk.Frame(sr, bg="#002855", padx=10, pady=5,
                         highlightbackground="#0A3A6A", highlightthickness=1)
            f.pack(side="left", padx=(0,6))
            tk.Label(f, text=label, bg="#002855", fg=SIDEBAR_TXT,
                     font=(FONT,7)).pack(anchor="w")
            tk.Label(f, text=str(value), bg="#002855", fg=colour,
                     font=(FONT,14,"bold")).pack(anchor="w")

        total     = len(client_orders)
        executed  = len([o for o in client_orders if o.get("status")=="EXECUTED"])
        pending   = len([o for o in client_orders if o.get("status") in ("PENDING","TAKEN","PARTIAL")])
        cancelled = len([o for o in client_orders if o.get("status")=="CANCELLED"])
        counters  = list(set(o.get("counter","") for o in client_orders if o.get("counter")))
        total_shares = sum(o.get("shares_executed",0) for o in client_orders)

        scard("Total Orders",    total,        FBC_ACCENT)
        scard("Executed",        executed,     "#4CAF50")
        scard("Open",            pending,      "#FBC02D")
        scard("Cancelled",       cancelled,    "#9E9E9E")
        scard("Counters Traded", len(counters),"#A855F7")
        scard("Shares Executed", f"{total_shares:,}", "#00A3E0")

        if not client_orders:
            tk.Label(self, text="No order history found for this client.",
                     bg=BG, fg="#8096B0", font=(FONT,11), pady=40).pack()
            tk.Button(self, text="Close", font=(FONT,10), bg=FBC_MID, fg=WHITE,
                      relief="flat", cursor="hand2", activebackground=FBC_DARK,
                      command=self.destroy, pady=5).pack(pady=10)
            return

        # ── Counter summary ──────────────────────────────────────────
        mid = tk.Frame(self, bg=BG); mid.pack(fill="x", padx=16, pady=(8,4))
        tk.Label(mid, text="Counters traded: " + "  ·  ".join(counters),
                 bg=BG, fg=FBC_DARK, font=(FONT,9,"bold")).pack(anchor="w")

        # ── Orders table ─────────────────────────────────────────────
        tk.Frame(self, bg=SEP_CLR, height=1).pack(fill="x", padx=16)

        tbl_outer = tk.Frame(self, bg=BG)
        tbl_outer.pack(fill="both", expand=True, padx=16, pady=(6,0))

        hdr_frame = tk.Frame(tbl_outer, bg=TBL_HDR_BG); hdr_frame.pack(fill="x")
        h_cols = [("Date",10),("Type",5),("Counter",10),("Shares",10),
                  ("Limit",9),("Status",13),("Filled",10),
                  ("Exec Price",10),("Dealer",13),("Entered By",11)]
        for label, w in h_cols:
            tk.Label(hdr_frame, text=label, bg=TBL_HDR_BG, fg=TBL_HDR_FG,
                     font=(FONT,8,"bold"), width=w, anchor="w",
                     padx=4, pady=5).pack(side="left")

        # Scrollable rows
        row_outer = tk.Frame(tbl_outer, bg=BG); row_outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(row_outer, bg=BG, highlightthickness=0)
        sb = tk.Scrollbar(row_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=BG)
        cid = canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(cid, width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-1*(e.delta//120),"units"))

        STATUS_CFG = {
            "PENDING":   ("#B45309","#FEF3C7"),
            "TAKEN":     ("#0066B3","#DBEAFE"),
            "EXECUTED":  ("#1A6B3A","#DCFCE7"),
            "PARTIAL":   ("#6B21A8","#F3E8FF"),
            "CANCELLED": ("#757575","#F5F5F5"),
        }

        for i, o in enumerate(client_orders):
            rbg = TBL_ROW_A if i % 2 == 0 else TBL_ROW_B
            status = o.get("status","")
            st_fg, st_bg = STATUS_CFG.get(status, ("#607080", BG))
            ot = o.get("order_type","")
            ot_bg = "#1A6B3A" if ot=="BUY" else "#B71C1C"

            row = tk.Frame(inner, bg=rbg); row.pack(fill="x")

            def lbl(text, w, fg="#1A2B3C", bg=rbg, bold=False, align="w"):
                _a = {"left": "w", "right": "e", "center": "center"}.get(align, align)
                tk.Label(row, text=str(text) if text else "—", bg=bg, fg=fg,
                         font=(FONT, 9, "bold" if bold else "normal"),
                         width=w, anchor=_a, padx=4, pady=5).pack(side="left")

            lbl(fmt_date_short(o.get("order_date","")), 10, fg="#607080")

            tk.Label(row, text=ot, bg=ot_bg, fg=WHITE,
                     font=(FONT,8,"bold"), width=5, anchor="center",
                     pady=5).pack(side="left")

            lbl(o.get("counter",""),        10, fg=FBC_DARK, bold=True)
            lbl(f"{o.get('num_shares',0):,}", 10, align="right")
            lbl(o.get("limit_price",""),    9)

            tk.Label(row, text=status, bg=st_bg, fg=st_fg,
                     font=(FONT,8,"bold"), width=13, anchor="w",
                     padx=4, pady=5).pack(side="left")

            filled = o.get("shares_executed",0)
            lbl(f"{filled:,}" if filled else "—", 10,
                fg="#1A6B3A" if filled else "#C0C8D8")
            lbl(o.get("execution_price",""), 10, fg="#607080")
            lbl(o.get("taken_by","") or "—", 13,
                fg="#C0392B" if o.get("taken_by") else "#C0C8D8")
            lbl(o.get("entered_by",""),      11, fg="#607080")

            tk.Frame(inner, bg=SEP_CLR, height=1).pack(fill="x")

        # ── Close ────────────────────────────────────────────────────
        btn_bar = tk.Frame(self, bg=BG, pady=8); btn_bar.pack(fill="x")
        tk.Button(btn_bar, text="  Close  ", font=(FONT,10), bg=FBC_MID, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_DARK,
                  command=self.destroy, pady=5).pack()
# ════════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self, dealer_name, settings):
        super().__init__()
        self.dealer_name = dealer_name; self.settings = settings
        self.title(f"FBC Order Manager  v{VERSION}  -  {dealer_name}")
        self.state("zoomed"); self.configure(bg=BG)
        url = settings.get("supa_url",""); key = settings.get("supa_key","")
        self.db = SheetsDB(url, key) if (url and key) else None
        self.orders = []
        self._filter     = "all"
        self._sort_col   = "entered_datetime"
        self._sort_asc   = False
        self._visible_orders     = []
        self._row_height         = 34
        self._render_buffer      = 5
        self._last_render_range  = None
        self._search_var = tk.StringVar()
        self._date_from  = tk.StringVar()
        self._date_to    = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._refresh())
        self._build()
        self._full_sync()
        self._schedule_auto_sync()

    def _build(self):
        top = tk.Frame(self, bg=FBC_DARK); top.pack(fill="x")
        tk.Frame(top, bg=FBC_ACCENT, width=6).pack(side="left", fill="y")
        tk.Label(top, text="FBC Securities - Order Manager", bg=FBC_DARK, fg=WHITE,
                 font=(FONT,13,"bold"), pady=10, padx=16).pack(side="left")
        self._sync_lbl = tk.Label(top, text="", bg=FBC_DARK, font=(FONT,9))
        self._sync_lbl.pack(side="left", padx=10)
        tk.Button(top, text="  Matched Trades  ", font=(FONT,10,"bold"),
                  bg="#6B21A8", fg=WHITE, relief="flat", cursor="hand2",
                  activebackground="#4C1D95", pady=5,
                  command=self._open_matched_trades).pack(side="right", padx=(0,6), pady=4)
        tk.Button(top, text="  📊 Report  ", font=(FONT, 10, "bold"),
                  bg="#1A6B3A", fg=WHITE, relief="flat", cursor="hand2",
                  activebackground="#145A32", pady=5,
                  command=self._open_report).pack(side="right", padx=6, pady=4)
        tk.Button(top, text="  New Order  ", font=(FONT,10,"bold"),
                  bg=FBC_ACCENT, fg=WHITE, relief="flat", cursor="hand2",
                  activebackground=FBC_MID, pady=5,
                  command=self._open_new_order).pack(side="right", padx=6, pady=4)
        tk.Button(top, text="DB Setup", font=(FONT,9), bg=FBC_DARK, fg=SIDEBAR_TXT,
                  relief="flat", cursor="hand2", activebackground=FBC_MID,
                  command=self._open_sheets_setup).pack(side="right", padx=4, pady=4)
        # Dealer avatar pill
        dealer_frame = tk.Frame(top, bg=FBC_DARK)
        dealer_frame.pack(side="right", padx=8)
        # Circle avatar with initial
        avatar = tk.Canvas(dealer_frame, width=26, height=26, bg=FBC_DARK, highlightthickness=0)
        avatar.pack(side="left", pady=6)
        avatar.create_oval(1, 1, 25, 25, fill=FBC_ACCENT, outline="")
        avatar.create_text(13, 13, text=self.dealer_name[0].upper(),
                           font=(FONT, 10, "bold"), fill=WHITE)
        tk.Label(dealer_frame, text=self.dealer_name, bg=FBC_DARK, fg=WHITE,
                 font=(FONT, 9, "bold")).pack(side="left", padx=(5, 0))
        tk.Label(top, text=f"v{VERSION}", bg=FBC_DARK, fg="#2A5A8A",
                 font=(FONT,9)).pack(side="right", padx=6)

        metric_bg = tk.Frame(self, bg=SIDEBAR_BG, pady=6); metric_bg.pack(fill="x")
        mr = tk.Frame(metric_bg, bg=SIDEBAR_BG); mr.pack(padx=20)
        def mcard(label, colour):
            f = tk.Frame(mr, bg="#002855", padx=12, pady=6,
                         highlightbackground="#0A3A6A", highlightthickness=1)
            f.pack(side="left", padx=(0,6))
            tk.Label(f, text=label, bg="#002855", fg=SIDEBAR_TXT, font=(FONT,7)).pack(anchor="w")
            v = tk.Label(f, text="0", bg="#002855", fg=colour, font=(FONT,16,"bold"))
            v.pack(anchor="w"); return v
        self._m_total     = mcard("Total",            FBC_ACCENT)
        self._m_pending   = mcard("Pending",          "#FBC02D")
        self._m_taken     = mcard("In Progress",      "#00A3E0")
        self._m_executed  = mcard("Executed Today",   "#4CAF50")
        self._m_partial   = mcard("Partial",          "#A855F7")
        self._m_overdue = mcard("Aging", "#FF9800")
        self._m_cancelled = mcard("Cancelled",        "#9E9E9E")
        # Make aging card clickable
        self._m_overdue.master.bind("<Button-1>", lambda e: self._open_aging())
        self._m_overdue.bind("<Button-1>", lambda e: self._open_aging())
        self._m_overdue.master.config(cursor="hand2")
        self._m_overdue.config(cursor="hand2")
        ctrl = tk.Frame(self, bg=BG); ctrl.pack(fill="x", padx=16, pady=(6,0))
        row1 = tk.Frame(ctrl, bg=BG); row1.pack(fill="x", pady=(0,3))
        sf = tk.Frame(row1, bg=BG); sf.pack(side="left", padx=(0,12))
        tk.Label(sf, text="Search:", bg=BG, font=(FONT,9)).pack(side="left")
        flat_entry(sf, self._search_var, width=22).pack(side="left", ipady=3, padx=(3,0))
        tk.Label(sf, text="Client / counter / CSD", bg=BG, fg="#A0B0C0",
                 font=(FONT,8)).pack(side="left", padx=(5,0))
        self._tabs = {}
        for key, label in [("all", "All"), ("pending", "Pending"), ("taken", "In Progress"),
                           ("executed", "Executed"), ("partial", "Partial"),
                           ("aging", "Aging"), ("cancelled", "Cancelled")]:
            b = tk.Button(row1, text=label, font=(FONT,9), relief="flat",
                          cursor="hand2", padx=8, pady=3,
                          command=lambda k=key: self._set_filter(k))
            b.pack(side="left", padx=(0,2)); self._tabs[key] = b
        self._paint_tabs()

        row2 = tk.Frame(ctrl, bg=BG); row2.pack(fill="x", pady=(0,4))
        tk.Label(row2, text="From", bg=BG, fg="#607080", font=(FONT,9)).pack(side="left")
        date_entry(row2, self._date_from, bg=BG).pack(side="left", padx=(3,0))
        tk.Label(row2, text="To", bg=BG, fg="#607080", font=(FONT,9)).pack(side="left", padx=(6,0))
        date_entry(row2, self._date_to, bg=BG).pack(side="left", padx=(3,0))
        tk.Button(row2, text="Apply", font=(FONT,8,"bold"), bg=FBC_MID, fg=WHITE,
                  relief="flat", cursor="hand2", activebackground=FBC_DARK, padx=7, pady=2,
                  command=self._refresh).pack(side="left", padx=(7,0))
        tk.Button(row2, text="Clear", font=(FONT,8), bg=BG, fg="#607080",
                  relief="flat", cursor="hand2", activebackground=SEP_CLR,
                  command=self._clear_dates).pack(side="left", padx=(3,0))
        for label, days in [("Today",0),("7d",7),("30d",30)]:
            tk.Button(row2, text=label, font=(FONT,8), bg=BG, fg=FBC_MID,
                      relief="flat", cursor="hand2", activebackground=SEP_CLR,
                      command=lambda d=days: self._quick_range(d)).pack(side="left", padx=(4,0))

        tk.Frame(self, bg=SEP_CLR, height=1).pack(fill="x", padx=16)

        tbl_outer = tk.Frame(self, bg=BG); tbl_outer.pack(fill="both", expand=True, padx=16, pady=6)
        self._columns = [
            ("status", "Status", 13, "w"), ("order_type", "Type", 5, "center"),
            ("counter", "Counter", 10, "w"), ("client_name", "Client", 14, "w"),
            ("csd_no", "CSD", 12, "w"),("qty_display", "Qty / Filled", 28, "e"),
            ("remaining_display", "Remaining", 20, "e"),
            ("limit_price", "Limit", 9, "e"), ("order_date", "Date", 10, "center"),
            ("entered_by", "Entered By", 10, "center"), ("dealer_col", "Dealer", 13, "w"),
            ("actions", "Actions", 18, "center"),
        ]
        hdr_frame = tk.Frame(tbl_outer, bg=TBL_HDR_BG); hdr_frame.pack(fill="x")
        self._hdr_btns = {}
        for col_id, label, width, anchor in self._columns:
            btn = tk.Button(hdr_frame, text=label, font=(FONT,8,"bold"),
                            bg=TBL_HDR_BG, fg=TBL_HDR_FG, relief="flat",
                            activebackground="#003070", activeforeground=WHITE,
                            cursor="hand2", anchor=anchor, width=width, pady=5,
                            command=lambda c=col_id: self._sort_by(c))
            btn.pack(side="left", padx=1); self._hdr_btns[col_id] = btn

        row_outer = tk.Frame(tbl_outer, bg=BG); row_outer.pack(fill="both", expand=True)
        self._tbl_canvas = tk.Canvas(row_outer, bg=BG, highlightthickness=0)
        tbl_sb = tk.Scrollbar(row_outer, orient="vertical", command=self._tbl_canvas.yview)
        self._tbl_canvas.configure(yscrollcommand=tbl_sb.set)
        tbl_sb.pack(side="right", fill="y")
        self._tbl_canvas.pack(side="left", fill="both", expand=True)
        self._tbl_inner = tk.Frame(self._tbl_canvas, bg=BG)
        self._tbl_inner_id = self._tbl_canvas.create_window((0,0), window=self._tbl_inner, anchor="nw")
        self._tbl_inner.bind("<Configure>",
                             lambda e: self._tbl_canvas.configure(scrollregion=self._tbl_canvas.bbox("all")))
        self._tbl_canvas.bind("<Configure>",
                              lambda e: self._tbl_canvas.itemconfig(self._tbl_inner_id, width=e.width))
        self._tbl_canvas.bind_all("<MouseWheel>",
                                  lambda e: (self._tbl_canvas.yview_scroll(-1*(e.delta//120),"units"),
                                             self._on_tbl_scroll()))
        self._tbl_canvas.bind("<Configure>", lambda e: self._on_tbl_scroll())

    def _sort_by(self, col_id):
        if self._sort_col == col_id: self._sort_asc = not self._sort_asc
        else: self._sort_col = col_id; self._sort_asc = True
        self._refresh()

    def _clear_dates(self):
        self._date_from.set(""); self._date_to.set(""); self._refresh()

    def _quick_range(self, days):
        from datetime import timedelta
        today = date.today()
        if days == 0:
            self._date_from.set(today.strftime("%d/%m/%Y")); self._date_to.set(today.strftime("%d/%m/%Y"))
        else:
            self._date_from.set((today-timedelta(days=days)).strftime("%d/%m/%Y"))
            self._date_to.set(today.strftime("%d/%m/%Y"))
        self._refresh()

    def _paint_tabs(self):
        for k, b in self._tabs.items():
            if k == self._filter: b.config(bg=FBC_DARK, fg=WHITE)
            else: b.config(bg=BG, fg="#607080", activebackground=SEP_CLR, activeforeground=FBC_DARK)

    def _set_filter(self, key):
        if key == "aging":
            self._open_aging()
            return
        self._filter = key;
        self._paint_tabs();
        self._refresh()
    def _update_sync_badge(self):
        if self.db and self.db.online: self._sync_lbl.config(text="Live", fg="#6EE7B7")
        elif self.db: self._sync_lbl.config(text="Offline (local cache)", fg="#FFB347")
        else: self._sync_lbl.config(text="No sync configured", fg="#607080")

    def _full_sync(self):
        def _do():
            orders = self.db.read_all() if self.db else load_orders_local()
            self.after(0, lambda: self._apply_orders(orders))
        threading.Thread(target=_do, daemon=True).start()

    def _apply_orders(self, orders):
        self.orders = orders; self._update_sync_badge(); self._refresh()

    def _schedule_auto_sync(self):
        self._full_sync(); self.after(20_000, self._schedule_auto_sync)

    # ── REFRESH ──────────────────────────────────────────────────────────
    def _refresh(self):
        orders = list(self.orders)
        search = self._search_var.get().strip().lower()
        today  = date.today().isoformat()
        d_from = parse_date_input(self._date_from.get())
        d_to   = parse_date_input(self._date_to.get())
        if d_from or d_to:
            filtered = []
            for o in orders:
                od_str = o.get("order_date","") or o.get("entered_datetime","")[:10]
                try: od = datetime.strptime(od_str[:10], "%Y-%m-%d").date()
                except: filtered.append(o); continue
                if d_from and od < d_from: continue
                if d_to   and od > d_to:   continue
                filtered.append(o)
            orders = filtered

        all_orders = self.orders
        def cnt(s): return len([o for o in all_orders if o["status"]==s])
        overdue_all = [o for o in all_orders if o["status"] in ("PENDING","TAKEN")
                       and days_since_str(o.get("entered_datetime","")) >= OVERDUE_DAYS]
        exec_today  = [o for o in all_orders if o["status"]=="EXECUTED"
                       and o.get("executed_datetime","")[:10]==today]
        self._m_total.config(    text=str(len(all_orders)))
        self._m_pending.config(  text=str(cnt("PENDING")))
        self._m_taken.config(    text=str(cnt("TAKEN")))
        self._m_executed.config( text=str(len(exec_today)))
        self._m_partial.config(  text=str(cnt("PARTIAL")))
        aging_pending = [o for o in all_orders if o["status"] == "PENDING"]
        aging_taken = [o for o in all_orders if o["status"] == "TAKEN"]
        aging_count = len(aging_pending) + len(aging_taken)
        self._m_overdue.config(text=str(aging_count))
        self._m_cancelled.config(text=str(cnt("CANCELLED")))

        overdue_set = {o["id"] for o in overdue_all}
        if   self._filter=="pending":   orders=[o for o in orders if o["status"]=="PENDING"]
        elif self._filter=="taken":     orders=[o for o in orders if o["status"]=="TAKEN"]
        elif self._filter=="executed":  orders=[o for o in orders if o["status"]=="EXECUTED"]
        elif self._filter=="partial":   orders=[o for o in orders if o["status"]=="PARTIAL"]
        elif self._filter == "aging":
            orders = [o for o in orders if o["status"] in ("PENDING", "TAKEN", "PARTIAL")]
        elif self._filter=="cancelled": orders=[o for o in orders if o["status"]=="CANCELLED"]

        if search:
            orders=[o for o in orders if
                    search in o.get("client_name","").lower() or
                    search in o.get("counter","").lower() or
                    search in o.get("csd_no","").lower() or
                    search in o.get("entered_by","").lower() or
                    search in o.get("id","").lower()]

        STATUS_ORDER = {"PENDING":0,"TAKEN":1,"PARTIAL":2,"EXECUTED":3,"CANCELLED":4}
        col = self._sort_col
        def sort_key(o):
            if col=="status": return (STATUS_ORDER.get(o["status"],5), o.get("entered_datetime",""))
            elif col in ("num_shares","shares_executed"): return o.get(col,0)
            elif col=="qty_display": return o.get("num_shares",0)
            elif col=="dealer_col": return (o.get("taken_by","") or "").lower()
            else: return str(o.get(col,"")).lower()
        orders = sorted(orders, key=sort_key, reverse=not self._sort_asc)

        for w in self._tbl_inner.winfo_children(): w.destroy()
        if not orders:
            msg = ("No orders yet.  Click New Order to post the first one."
                   if self._filter=="all" and not search and not d_from and not d_to
                   else "No orders match the current filters.")
            tk.Label(self._tbl_inner, text=msg, bg=BG, fg="#8096B0",
                     font=(FONT,11), pady=40).pack()
            self._visible_orders = []; return

        self._visible_orders = orders
        self._last_render_range = None
        total_h = len(orders) * self._row_height
        self._tbl_canvas.configure(scrollregion=(0,0,0,total_h+40))
        self._tbl_canvas.yview_moveto(0)
        self._on_tbl_scroll()

    # ── VIRTUAL SCROLL ────────────────────────────────────────────────────
    def _on_tbl_scroll(self, event=None):
        if not self._visible_orders: return
        canvas_h = self._tbl_canvas.winfo_height()
        if canvas_h <= 1:
            self.after(50, self._on_tbl_scroll); return
        try: top_frac = self._tbl_canvas.yview()[0]
        except: top_frac = 0.0
        total_h = len(self._visible_orders) * self._row_height
        top_px  = top_frac * total_h
        first = max(0, int(top_px / self._row_height) - self._render_buffer)
        last  = min(len(self._visible_orders),
                    int((top_px+canvas_h) / self._row_height) + self._render_buffer + 1)
        if self._last_render_range == (first, last): return
        self._last_render_range = (first, last)
        for w in self._tbl_inner.winfo_children(): w.destroy()
        if first > 0:
            tk.Frame(self._tbl_inner, bg=BG, height=first*self._row_height).pack(fill="x")
        for i in range(first, last):
            self._table_row(self._tbl_inner, self._visible_orders[i], row_index=i)
        remaining = len(self._visible_orders) - last
        if remaining > 0:
            tk.Frame(self._tbl_inner, bg=BG, height=remaining*self._row_height).pack(fill="x")
        self._tbl_inner.update_idletasks()

    # ── TABLE ROW ─────────────────────────────────────────────────────────
    def _table_row(self, parent, o, row_index):
        status = o["status"]
        is_overdue = (status in ("PENDING","TAKEN") and
                      days_since_str(o.get("entered_datetime","")) >= OVERDUE_DAYS)
        row_bg = TBL_ROW_A if row_index % 2 == 0 else TBL_ROW_B
        if is_overdue: row_bg = "#FFF5F5"
        accent_colours = {"PENDING":"#FBC02D","TAKEN":"#00A3E0","EXECUTED":"#4CAF50",
                          "PARTIAL":"#A855F7","CANCELLED":"#9E9E9E"}
        if is_overdue: accent_colours["PENDING"] = "#B71C1C"; accent_colours["TAKEN"] = "#B71C1C"
        accent = accent_colours.get(status, "#9E9E9E")
        row = tk.Frame(parent, bg=row_bg); row.pack(fill="x")
        tk.Frame(row, bg=accent, width=4).pack(side="left", fill="y")

        def lbl(text, w, anchor="w", fg="#1A2B3C", bold=False, bg=row_bg):
            tk.Label(row, text=text, bg=bg, fg=fg,
                     font=(FONT,9,"bold" if bold else "normal"),
                     width=w, anchor=anchor, padx=2, pady=6).pack(side="left")

        status_cfg = {
            "PENDING":   ("PENDING",     "#B45309","#FEF3C7"),
            "TAKEN":     ("IN PROGRESS", "#0066B3","#DBEAFE"),
            "EXECUTED":  ("EXECUTED",    "#1A6B3A","#DCFCE7"),
            "PARTIAL":   ("PARTIAL",     "#6B21A8","#F3E8FF"),
            "CANCELLED": ("CANCELLED",   "#757575","#F5F5F5"),
        }
        # REPLACE WITH:
        is_over_exec = "OVER-EXECUTION" in (o.get("notes", "") or "")
        if is_over_exec:
            st_txt, st_fg, st_bg = "🚨 OVER-EXEC", "#B71C1C", "#FFE4E4"
        elif is_overdue:
            st_txt, st_fg, st_bg = "OVERDUE", "#B71C1C", "#FEE2E2"
        else:
            st_txt, st_fg, st_bg = status_cfg.get(status, (status, "#607080", BG))
        tk.Label(row, text=st_txt, bg=st_bg, fg=st_fg,
                 font=(FONT,8,"bold"), width=14, anchor="w",
                 padx=4, pady=6).pack(side="left", padx=(0,2))

        ot = o["order_type"]; ot_bg = "#1A6B3A" if ot=="BUY" else "#B71C1C"
        tk.Label(row, text=ot, bg=ot_bg, fg=WHITE,
                 font=(FONT,8,"bold"), width=5, anchor="center", pady=6).pack(side="left", padx=(0,2))

        lbl(o.get("counter",""), 10, bold=True, fg=FBC_DARK)
        client = o.get("client_name", "")
        if len(client) > 14: client = client[:13] + "..."
        lbl(client, 14)
        lbl(o.get("csd_no", ""), 12, fg="#607080")
        filled = o.get("shares_executed", 0);
        ordered = o.get("num_shares", 0)
        mode = o.get("amount_mode", "SHARES")
        exch_col = o.get("exchange", "ZSE")
        currency_col = "USD" if exch_col == "VFEX" else "ZiG"
        total_amount_col = o.get("total_amount", "")
        is_over_exec = "OVER-EXECUTION" in (o.get("notes", "") or "")
        rem_txt = "—"
        rem_fg = "#C0C8D8"
        if mode == "AMOUNT":
            try:
                amt_f = float(str(total_amount_col).replace(",", ""))
            except:
                amt_f = 0.0
            if filled and amt_f > 0:
                try:
                    ep = float(str(o.get("execution_price", "")).replace(",", ""))
                    amt_filled = filled * ep
                    pct = int((amt_filled / amt_f) * 100)
                    qty_txt = f"{amt_filled:,.2f} / {amt_f:,.2f} ({pct}%)"
                    amt_remaining = max(0.0, amt_f - amt_filled)
                    if amt_remaining > 0.01:
                        rem_txt = f"{amt_remaining:,.2f}"
                        rem_fg = "#B45309"
                    else:
                        rem_txt = "✓ FULL"
                        rem_fg = "#1A6B3A"
                except:
                    qty_txt = f"{amt_f:,.2f}"
            else:
                qty_txt = f"{amt_f:,.2f}" if amt_f else str(total_amount_col)

        elif filled and ordered:
            remaining = ordered - filled
            pct = int((filled / ordered) * 100) if ordered else 0
            qty_txt = f"{filled:,} / {ordered:,} ({pct}%)"
            if remaining > 0:
                rem_txt = f"{remaining:,} sh"
                rem_fg = "#B45309"
            else:
                rem_txt = "✓ FULL"
                rem_fg = "#1A6B3A"
        else:
            qty_txt = f"{ordered:,}" if ordered else "-"
            rem_txt = "—";
            rem_fg = "#C0C8D8"

        qty_fg = "#B71C1C" if is_over_exec else (
            "#0066B3" if mode == "AMOUNT" else
            "#6B21A8" if (status == "PARTIAL" and filled) else "#1A2B3C")
        lbl(qty_txt, 28, anchor="e", fg=qty_fg, bold=(mode == "AMOUNT" or status == "PARTIAL" or is_over_exec))
        lbl(rem_txt, 20, anchor="e", fg=rem_fg, bold=(rem_txt not in ("—", "✓ FULL") or rem_txt == "✓ FULL"))
        lbl(o.get("limit_price","-"), 9, anchor="e", fg="#607080")
        lbl(fmt_date_short(o.get("order_date","")), 10, anchor="center", fg="#607080")
        lbl(o.get("entered_by", ""), 10, anchor="center", fg="#607080")
        # Dealer column - RED for TAKEN/PARTIAL
        taken   = o.get("taken_by","")
        exec_by = o.get("executed_by","")
        if taken and status in ("TAKEN","PARTIAL"):
            dealer_txt = f"  {taken}"; dealer_fg = "#C0392B"; dealer_bg = "#FDECEA"
        elif taken and status == "EXECUTED":
            dealer_txt = taken; dealer_fg = "#607080"; dealer_bg = row_bg
        elif exec_by:
            short_exec = exec_by.replace("AUTO - ","").replace(" (matched trades)","")
            if len(short_exec) > 14: short_exec = short_exec[:13]+"..."
            dealer_txt = short_exec; dealer_fg = "#607080"; dealer_bg = row_bg
        else:
            dealer_txt = "-"; dealer_fg = "#C0C8D8"; dealer_bg = row_bg
        tk.Label(row, text=dealer_txt, bg=dealer_bg, fg=dealer_fg,
                 font=(FONT, 8, "bold" if (taken and status in ("TAKEN", "PARTIAL")) else "normal"),
                 width=13, anchor="w", padx=4, pady=6).pack(side="left", padx=(0, 2))

        act_frame = tk.Frame(row, bg=row_bg); act_frame.pack(side="left", padx=4)
        def mini_btn(text, bg, fg, cmd, width=5):
            return tk.Button(act_frame, text=text, font=(FONT,8), bg=bg, fg=fg,
                             relief="flat", cursor="hand2", padx=3, pady=2, width=width,
                             activebackground=fg, activeforeground=WHITE, command=cmd)

        mini_btn("Detail", BG, FBC_MID,
                 lambda ord=o: OrderDetailDialog(self, ord)).pack(side="left", padx=1)
        mini_btn("PDF", BG, "#6B21A8",
                 lambda ord=o: self._print_slip(ord), width=3).pack(side="left", padx=1)
        mini_btn("Hist", BG, "#B45309",
                 lambda ord=o: ClientHistoryDialog(
                     self, ord.get("client_name", ""),
                     ord.get("csd_no", ""), self.orders),
                 width=4).pack(side="left", padx=1)

        if status == "PENDING":
            mini_btn("Edit", BG, "#607080",
                     lambda ord=o: self._open_edit(ord), width=4).pack(side="left", padx=1)
            mini_btn("Take", FBC_MID, WHITE,
                     lambda ord=o: self._take_order(ord), width=4).pack(side="left", padx=1)
            mini_btn("X", "#FEE2E2", "#B71C1C",
                     lambda ord=o: self._open_cancel(ord), width=2).pack(side="left", padx=1)

        elif status == "TAKEN":
            if o.get("taken_by") == self.dealer_name:
                mini_btn("Execute", "#1A6B3A", WHITE,
                         lambda ord=o: self._open_execute(ord), width=6).pack(side="left", padx=1)
                mini_btn("Release", BG, "#607080",
                         lambda ord=o: self._release_order(ord), width=5).pack(side="left", padx=1)
                mini_btn("X", "#FEE2E2", "#B71C1C",
                         lambda ord=o: self._open_cancel(ord), width=2).pack(side="left", padx=1)
            else:
                tk.Label(act_frame, text=f"Held: {o.get('taken_by','')}",
                         bg=row_bg, fg="#C0392B", font=(FONT,8,"bold")).pack(side="left")

        elif status == "PARTIAL":
            can_act = (o.get("taken_by")==self.dealer_name or not o.get("taken_by"))
            if can_act:
                mini_btn("Execute", "#6B21A8", WHITE,
                         lambda ord=o: self._open_execute(ord), width=6).pack(side="left", padx=1)
                mini_btn("X", "#FEE2E2", "#B71C1C",
                         lambda ord=o: self._open_cancel(ord), width=2).pack(side="left", padx=1)
            else:
                tk.Label(act_frame, text=f"Held: {o.get('taken_by','')}",
                         bg=row_bg, fg="#C0392B", font=(FONT,8,"bold")).pack(side="left")

        elif status in ("CANCELLED","EXECUTED"):
            mini_btn("Del", BG, "#CBD5E1",
                     lambda ord=o: self._delete(ord), width=3).pack(side="left", padx=1)

        tk.Frame(parent, bg=SEP_CLR, height=1).pack(fill="x")

    # ── ORDER ACTIONS ─────────────────────────────────────────────────────
    def _take_order(self, order):
        if order["status"] != "PENDING":
            messagebox.showinfo("Already taken","This order has already been taken.",parent=self); return
        if not messagebox.askyesno("Take Order",
            f"Claim this order?\n\n{order['order_type']}  {order['counter']}\n"
            f"Client: {order['client_name']}\n\nYour name ({self.dealer_name}) will be stamped.",
            parent=self): return
        order["status"]="TAKEN"; order["taken_by"]=self.dealer_name
        order["taken_datetime"]=datetime.now().isoformat(); self._persist(order)

    def _release_order(self, order):
        if not messagebox.askyesno("Release Order",
            "Release this order back to PENDING?\n\nOther dealers will be able to take it.",
            parent=self): return
        order["status"]="PENDING"; order["taken_by"]=""; order["taken_datetime"]=""
        self._persist(order)

    def _open_execute(self, order):
        ExecuteOrderDialog(self, order, self.dealer_name, self._on_executed)

    def _on_executed(self, order):
        self._persist(order)

    def _open_cancel(self, order):
        CancelOrderDialog(self, order, self.dealer_name, self._on_cancelled)

    def _on_cancelled(self, order):
        self._persist(order)

    def _open_edit(self, order):
        EditOrderDialog(self, order, self._on_edited)

    def _on_edited(self, order):
        self._persist(order)

    def _open_new_order(self):
        NewOrderDialog(self, self.dealer_name, self._add_new, orders=self.orders)

    def _open_matched_trades(self):
        MatchedTradesDialog(self, self.orders, self.dealer_name, self._on_bulk_execute)

    def _on_bulk_execute(self, updated_orders):
        def _do():
            for order in updated_orders:
                if self.db: self.db.update_order(order)
            save_orders_local(self.orders); self.after(0, self._refresh)
        threading.Thread(target=_do, daemon=True).start()

    def _add_new(self, order):
        def _do():
            if self.db: self.db.append_order(order)
            self.orders.append(order); save_orders_local(self.orders)
            self.after(0, self._refresh)
        threading.Thread(target=_do, daemon=True).start()

    def _persist(self, order):
        def _do():
            if self.db: self.db.update_order(order)
            save_orders_local(self.orders); self.after(0, self._refresh)
        threading.Thread(target=_do, daemon=True).start()

    def _delete(self, order):
        if not messagebox.askyesno("Delete",
            f"Permanently delete the record for {order['client_name']}?\n\nThis cannot be undone.",
            parent=self): return
        def _do():
            if self.db: self.db.delete_order(order["id"])
            if order in self.orders: self.orders.remove(order)
            save_orders_local(self.orders); self.after(0, self._refresh)
        threading.Thread(target=_do, daemon=True).start()

    def _open_aging(self):
        AgingDialog(self, self.orders, self.dealer_name)

    def _open_report(self):
        DailyReportDialog(self, self.orders)
    def _print_slip(self, order):
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        save_dir = downloads if os.path.isdir(downloads) else os.path.expanduser("~")
        filename = f"FBC_Order_{order['id']}_{order.get('client_name', '').replace(' ', '_')}.pdf"
        save_path = os.path.join(save_dir, filename)
        ok, err = OrderSlipPDF.generate(order, save_path)
        if ok:
            messagebox.showinfo("PDF Saved",
                                f"Order slip saved to:\n{save_path}\n\nOpening now...",
                                parent=self)
            try:
                os.startfile(save_path)
            except:
                pass
        else:
            messagebox.showerror("PDF Error", f"Could not generate PDF:\n{err}",
                                 parent=self)
    def _open_sheets_setup(self):
        SheetsSetupDialog(self, self.settings, self._on_sheets_saved)

    def _on_sheets_saved(self, url, key):
        if url and key:
            self.settings["supa_url"]=url; self.settings["supa_key"]=key
            self.db=SheetsDB(url,key); self._full_sync()
        self._update_sync_badge()


if __name__ == "__main__":
    _tmp_root = tk.Tk(); _tmp_root.withdraw()
    FONT = _pick_font(); _tmp_root.destroy()
    check_and_apply_update()
    login = LoginDialog(); login.mainloop()
    if not login.authenticated: sys.exit(0)
    settings = load_settings()
    if not settings.get("supa_url") or not settings.get("supa_key"):
        _tmp = tk.Tk(); _tmp.withdraw()
        _saved = {}
        def _on_setup(url, key): _saved["supa_url"]=url; _saved["supa_key"]=key
        dlg = SheetsSetupDialog(_tmp, settings, _on_setup)
        _tmp.wait_window(dlg); _tmp.destroy()
        if _saved: settings.update(_saved); save_settings(settings)
    settings = load_settings()
    app = App(dealer_name=login.dealer_name, settings=settings)
    app.mainloop()
